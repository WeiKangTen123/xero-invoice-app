import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

REPORT_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report"
ARCH_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report/arch_recommendation"

# ==============================================================================
# 1. EXCEL REPORT GENERATOR (With Setup Labor, IMAP & Gemini Registration)
# ==============================================================================
def create_excel_report(out_path):
    wb = openpyxl.Workbook()
    
    NAVY_DARK = "1E3A8A"
    NAVY_LIGHT = "DBEAFE"
    SLATE_DARK = "334155"
    SLATE_LIGHT = "F8FAFC"
    GREEN_ACCENT = "0D9488"
    GREEN_LIGHT = "CCFBF1"
    GRAY_TEXT = "64748B"
    WHITE = "FFFFFF"
    
    font_title = Font(name="Calibri", size=16, bold=True, color=NAVY_DARK)
    font_subtitle = Font(name="Calibri", size=11, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=12, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=11, color="1E293B")
    font_data_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_total = Font(name="Calibri", size=11, bold=True, color=NAVY_DARK)
    
    fill_sec_hdr = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=SLATE_DARK, end_color=SLATE_DARK, fill_type="solid")
    fill_alt_row = PatternFill(start_color=SLATE_LIGHT, end_color=SLATE_LIGHT, fill_type="solid")
    fill_highlight = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    fill_total = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    total_border = Border(
        top=Side(style='thin', color='1E3A8A'),
        bottom=Side(style='double', color='1E3A8A'),
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    # -------------------------------------------------------------
    # TAB 1: EXECUTIVE COMMERCIAL PROPOSAL & SETUP FEES
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Client Pricing & Setup Model"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "XERO INVOICE AUTOMATION — CLIENT COST & ONBOARDING PROPOSAL"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_left
    
    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Complete Commercial Model: One-Time Onboarding Fees, IMAP & Gemini Setup, and Monthly Retainers"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = align_left
    
    # Metadata
    ws1["A4"] = "Target Hardware:"
    ws1["B4"] = "8 GB RAM | 160-200 GB NVMe SSD | 4 vCPU"
    ws1["A5"] = "Target Architecture:"
    ws1["B5"] = "Multi-Tenant Real-Time IMAP + Xero OAuth + Bull"
    ws1["D4"] = "Prepared Date:"
    ws1["E4"] = "August 2026"
    ws1["D5"] = "Base Currency:"
    ws1["E5"] = "USD ($)"
    for r in [4, 5]:
        ws1[f"A{r}"].font = font_data_bold
        ws1[f"B{r}"].font = font_data
        ws1[f"D{r}"].font = font_data_bold
        ws1[f"E{r}"].font = font_data

    # Section 1: Client Pricing Packages
    ws1.merge_cells("A7:G7")
    ws1["A7"] = "RECOMMENDED CLIENT CHARGING PACKAGES (HOW TO PRICE TO CLIENTS)"
    ws1["A7"].font = font_sec_hdr
    ws1["A7"].fill = fill_sec_hdr

    headers_pricing = [
        "Package Tier", "One-Time Setup Fee", "Monthly Retainer", 
        "Included Invoices", "Included Setup Services", "Net Monthly Profit", "Target Client Profile"
    ]
    for col_num, h in enumerate(headers_pricing, 1):
        cell = ws1.cell(row=8, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    pricing_data = [
        ["Standard Business Tier (Recommended)", 1200.00, 199.00, "Up to 500 inv/mo", "Full IMAP, Xero OAuth, Gemini Setup & 1hr Training", 185.10, "SME with 50–500 monthly bills"],
        ["Growth / High-Volume Tier", 1800.00, 349.00, "Up to 1,500 inv/mo", "Custom chart of accounts, multi-currency & priority queue", 333.50, "Trading, logistics, e-commerce with high volume"],
        ["Accounting Agency / Multi-Entity Tier", 2800.00, 699.00, "Up to 3,500 inv/mo", "Multi-tenant portal setup (up to 10 company sub-accounts)", 675.00, "Accounting firms managing multiple Xero books"],
    ]

    for row_idx, row_vals in enumerate(pricing_data, start=9):
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [2, 3, 6]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
                if col_idx == 6:
                    cell.fill = fill_highlight
                    cell.font = font_total
            else:
                cell.font = font_data
                cell.alignment = align_center if col_idx == 4 else align_left

    # Section 2: Onboarding Setup Labor & Direct Registration Costs
    ws1.merge_cells("A13:G13")
    ws1["A13"] = "ITEMIZED ONBOARDING & SETUP COSTS (PER NEW CLIENT COMPANY)"
    ws1["A13"].font = font_sec_hdr
    ws1["A13"].fill = fill_sec_hdr

    headers_setup = ["Onboarding Activity / Setup Step", "Estimated Time", "Direct Vendor Fee", "Labor Value (@ $60/hr)", "Billable to Client", "Deliverables & Verification"]
    for col_num, h in enumerate(headers_setup, 1):
        cell = ws1.cell(row=14, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    setup_data = [
        ["1. Xero OAuth2 / Custom Connection Setup", "1.5 hours", 0.00, 90.00, 250.00, "Xero Dev App, tenant authorization, account code 310 & tax mapping"],
        ["2. IMAP Email Mailbox & Security Configuration", "1.0 hour", 0.00, 60.00, 200.00, "Gmail/M365 App Password, 2FA setup, 24/7 IMAP IDLE listener test"],
        ["3. Google Gemini AI API Registration & Key Setup", "0.5 hours", 0.00, 30.00, 150.00, "AI Studio key registration, 15 RPM rate limiting, key rotation pool"],
        ["4. Sample Pilot Run & Extraction Tuning", "1.5 hours", 0.00, 90.00, 350.00, "15–20 test vendor PDFs, tax validation, PayNow/Bank reference check"],
        ["5. Client Finance Staff Handover & Training", "1.0 hour", 0.00, 60.00, 250.00, "1-on-1 walkthrough of review dashboard, PDF preview, manual rescan"],
    ]

    for row_idx, row_vals in enumerate(setup_data, start=15):
        fill_to_use = fill_alt_row if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [3, 4, 5]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            else:
                cell.font = font_data
                cell.alignment = align_center if col_idx == 2 else align_left

    tot_setup_row = len(setup_data) + 15
    ws1.cell(row=tot_setup_row, column=1, value="TOTAL ONBOARDING & SETUP FEE").font = font_total
    ws1.cell(row=tot_setup_row, column=2, value="5.5 Hours").font = font_total
    ws1.cell(row=tot_setup_row, column=2).alignment = align_center
    
    ws1.cell(row=tot_setup_row, column=3, value=f"=SUM(C15:C{tot_setup_row-1})").number_format = "$#,##0.00"
    ws1.cell(row=tot_setup_row, column=4, value=f"=SUM(D15:D{tot_setup_row-1})").number_format = "$#,##0.00"
    ws1.cell(row=tot_setup_row, column=5, value=f"=SUM(E15:E{tot_setup_row-1})").number_format = "$#,##0.00"
    
    for c in range(1, 7):
        ws1.cell(row=tot_setup_row, column=c).fill = fill_total
        ws1.cell(row=tot_setup_row, column=c).border = total_border
        if c in [3, 4, 5]:
            ws1.cell(row=tot_setup_row, column=c).font = font_total
            ws1.cell(row=tot_setup_row, column=c).alignment = align_right

    # -------------------------------------------------------------
    # TAB 2: ITEMIZED INFRASTRUCTURE & VENDOR EXPENSES
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Itemized Operating Costs")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "ITEMIZED INFRASTRUCTURE & THIRD-PARTY EXPENSES (8GB / 200GB SERVER)"
    ws2["A1"].font = font_title
    
    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Audit of direct monthly cloud bills, API costs, email mailboxes & backup expenses"
    ws2["A2"].font = font_subtitle
    
    headers_infra = ["Category", "Line Item / Provider", "Base Specs / Quota", "Pricing Unit", "DigitalOcean Route", "GCP Route"]
    for c_idx, h in enumerate(headers_infra, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    infra_items = [
        ["Cloud Server (VPS)", "4 vCPU / 8 GB RAM / 160 GB NVMe SSD", "Handles 20-50 active IMAP sockets", "Fixed Monthly", 48.00, 68.50],
        ["NVMe Storage Volume", "+40 GB block storage (200GB Total SSD)", "Holds 400,000+ invoice PDFs & DB", "Fixed ($0.10/GB)", 4.00, 8.00],
        ["Dedicated Static IPv4", "1 Public IP for Xero Webhooks/SSH", "Fixed IP address", "Monthly", 0.00, 3.65],
        ["Bandwidth / Transfer", "High-speed data egress for PDFs & UI", "5,000 GB (5TB) DO / Pay-per-GB GCP", "Included / Usage", 0.00, 8.00],
        ["AI LLM Parsing (Gemini)", "Google Gemini Flash-Lite (5,000 inv/mo)", "5 Million tokens/mo @ $0.075/1M", "Pay-as-you-go", 0.75, 0.75],
        ["Gemini API Registration", "Google AI Studio Account Creation", "Zero upfront registration fee", "Free Registration", 0.00, 0.00],
        ["Xero Platform API", "OAuth2 / Custom Connection (60 RPM/org)", "Multi-tenant developer portal", "Free Developer Tier", 0.00, 0.00],
        ["Email Intake (IMAP)", "Client IMAP / Gmail App Password", "Direct connection to existing inbox", "Existing Account", 0.00, 0.00],
        ["Domain, DNS & SSL", "Custom Domain Name (.com / .app)", "Automated Let's Encrypt SSL", "Amortized ($15/yr)", 1.25, 1.25],
        ["Automated Backups", "Daily SQLite DB snapshots (Cloudflare R2/S3)", "Automated disaster recovery", "Storage ($0.015/GB)", 2.00, 5.20],
        ["Monitoring & Alerts", "Slack Webhook alerts + Uptime Kuma Check", "Instant crash/fail alerts", "Free Tier", 0.00, 0.00]
    ]

    for r_idx, row_vals in enumerate(infra_items, start=5):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx in [1, 2]:
                cell.font = font_data_bold if c_idx == 1 else font_data
                cell.alignment = align_left
            elif c_idx in [3, 4]:
                cell.font = font_data
                cell.alignment = align_center
            elif c_idx in [5, 6]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"

    tot_infra_row = len(infra_items) + 5
    ws2.cell(row=tot_infra_row, column=1, value="TOTAL ESTIMATED DIRECT MONTHLY INFRASTRUCTURE COST").font = font_total
    ws2.merge_cells(start_row=tot_infra_row, start_column=1, end_row=tot_infra_row, end_column=4)
    for c in range(1, 5):
        ws2.cell(row=tot_infra_row, column=c).fill = fill_total
        ws2.cell(row=tot_infra_row, column=c).border = total_border

    c_do_tot = ws2.cell(row=tot_infra_row, column=5, value=f"=SUM(E5:E{tot_infra_row-1})")
    c_do_tot.font = font_total
    c_do_tot.fill = fill_total
    c_do_tot.border = total_border
    c_do_tot.alignment = align_right
    c_do_tot.number_format = "$#,##0.00"

    c_gcp_tot = ws2.cell(row=tot_infra_row, column=6, value=f"=SUM(F5:F{tot_infra_row-1})")
    c_gcp_tot.font = font_total
    c_gcp_tot.fill = fill_total
    c_gcp_tot.border = total_border
    c_gcp_tot.alignment = align_right
    c_gcp_tot.number_format = "$#,##0.00"

    # -------------------------------------------------------------
    # TAB 3: MULTI-COMPANY SAAS PROFITABILITY & ROI
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="SaaS Profit & Client ROI")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "MULTI-COMPANY SAAS MARGIN MODEL & CLIENT ROI JUSTIFICATION"
    ws3["A1"].font = font_title
    
    headers_scale = ["Scale Stage", "Active Companies", "Monthly Retainer / Co.", "Gross Monthly Revenue", "Total Server Cost", "Total AI & Misc Costs", "Net Monthly Profit"]
    for col_num, h in enumerate(headers_scale, 1):
        cell = ws3.cell(row=4, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    scale_data = [
        ["Phase 1: Initial Pool", 5, 199.00, "=B5*C5", 56.00, 5.00, "=D5-E5-F5"],
        ["Phase 1: Expansion", 15, 199.00, "=B6*C6", 56.00, 15.00, "=D6-E6-F6"],
        ["Phase 2: Decoupled 3-Tier", 35, 199.00, "=B7*C7", 135.00, 35.00, "=D7-E7-F7"],
        ["Phase 2: Scaled Multi-Tenant", 75, 199.00, "=B8*C8", 145.00, 75.00, "=D8-E8-F8"],
    ]

    for row_idx, row_vals in enumerate(scale_data, start=5):
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [3, 4, 5, 6, 7]:
                cell.font = font_total if col_idx == 7 else font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
                if col_idx == 7:
                    cell.fill = fill_highlight
            else:
                cell.font = font_data
                cell.alignment = align_center

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str += "    "
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
    wb.save(out_path)
    print(f"Excel report successfully created at: {out_path}")


# ==============================================================================
# 2. PDF REPORT GENERATOR (With Setup Labor, IMAP & Gemini Registration)
# ==============================================================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Header (pages 2+)
        if self._pageNumber > 1:
            self.drawString(54, 11 * inch - 36, "XERO INVOICE AUTOMATION — COST & SETUP PROPOSAL")
            self.drawRightString(8.5 * inch - 54, 11 * inch - 36, "CONFIDENTIAL")
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(54, 11 * inch - 42, 8.5 * inch - 54, 11 * inch - 42)
            
        # Footer
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(54, 46, 8.5 * inch - 54, 46)
        
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawString(54, 32, "Commercial Cost Estimation & Onboarding Proposal | August 2026")
        self.drawRightString(8.5 * inch - 54, 32, page_str)
        self.restoreState()


def create_pdf_report(out_path):
    doc = SimpleDocTemplate(
        out_path,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    c_navy = colors.HexColor("#1E3A8A")
    c_slate = colors.HexColor("#334155")
    c_body = colors.HexColor("#1E293B")
    c_teal = colors.HexColor("#0D9488")
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=17, leading=21,
        textColor=c_navy, spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'],
        fontName='Helvetica-Oblique', fontSize=9, leading=12,
        textColor=colors.HexColor("#64748B"), spaceAfter=10
    )
    h1_style = ParagraphStyle(
        'SecH1', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10.5, leading=14,
        textColor=c_navy, spaceBefore=9, spaceAfter=4, keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyDark', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, leading=11,
        textColor=c_body, spaceAfter=4
    )
    th_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
        textColor=colors.white, alignment=1
    )
    td_left = ParagraphStyle(
        'TableCellLeft', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_bold = ParagraphStyle(
        'TableCellBold', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_right = ParagraphStyle(
        'TableCellRight', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=2
    )
    td_center = ParagraphStyle(
        'TableCellCenter', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=1
    )
    
    elements = []
    
    # ── HEADER BLOCK ──
    elements.append(Paragraph("XERO INVOICE AUTOMATION — COST & ONBOARDING PROPOSAL", title_style))
    elements.append(Paragraph("Itemized Setup Labor, IMAP & Gemini Registration, Real-Time Architecture & Monthly Retainers", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=c_navy, spaceAfter=8))
    
    # ── METADATA TABLE ──
    meta_table_data = [
        [Paragraph("<b>Target Stack:</b> Multi-Tenant Node.js + SQLite WAL + React + Bull", td_left), Paragraph("<b>Date:</b> August 2026", td_left)],
        [Paragraph("<b>Production Hardware:</b> 8GB RAM | 160–200GB NVMe SSD | 4 vCPU", td_left), Paragraph("<b>Server Cost:</b> ~$56.00 / mo on DigitalOcean", td_left)],
        [Paragraph("<b>Setup Scope:</b> Xero OAuth + IMAP + Gemini API + Training", td_left), Paragraph("<b>Recommended Retainer:</b> $199.00 / company / month", td_left)]
    ]
    t_meta = Table(meta_table_data, colWidths=[270, 234])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 6))
    
    # ── SECTION 1: ONBOARDING & SETUP FEES ──
    elements.append(Paragraph("1. ITEMIZED ONBOARDING & SETUP FEE BREAKDOWN (HELPING CLIENTS SET UP)", h1_style))
    elements.append(Paragraph(
        "Setting up a new client involves configuring Xero API permissions, email inbox authentication, AI keys, testing, and staff training:",
        body_style
    ))
    
    setup_pdf_data = [
        [
            Paragraph("Setup & Onboarding Activity", th_style),
            Paragraph("Est. Time", th_style),
            Paragraph("Vendor Fee", th_style),
            Paragraph("Labor Value", th_style),
            Paragraph("Client Billable", th_style),
            Paragraph("Deliverables Included", th_style),
        ],
        [
            Paragraph("<b>1. Xero Integration Setup</b>", td_left),
            Paragraph("1.5 hrs", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("$90.00", td_right),
            Paragraph("<b>$250.00</b>", td_right),
            Paragraph("OAuth2/Custom Connection, tenant ID linking, account code 310 & tax mapping", td_left),
        ],
        [
            Paragraph("<b>2. IMAP Email Pipeline Setup</b>", td_left),
            Paragraph("1.0 hr", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("$60.00", td_right),
            Paragraph("<b>$200.00</b>", td_right),
            Paragraph("Gmail/M365 App Password, 2FA security, 24/7 IMAP IDLE real-time test", td_left),
        ],
        [
            Paragraph("<b>3. Gemini AI API Registration</b>", td_left),
            Paragraph("0.5 hrs", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("$30.00", td_right),
            Paragraph("<b>$150.00</b>", td_right),
            Paragraph("Google AI Studio account, API key generation, key rotation & 15 RPM throttling", td_left),
        ],
        [
            Paragraph("<b>4. Pilot Run & Prompt Tuning</b>", td_left),
            Paragraph("1.5 hrs", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("$90.00", td_right),
            Paragraph("<b>$350.00</b>", td_right),
            Paragraph("Testing 15–20 real vendor PDFs, line item precision, PayNow/Bank parsing check", td_left),
        ],
        [
            Paragraph("<b>5. Staff Training & Handover</b>", td_left),
            Paragraph("1.0 hr", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("$60.00", td_right),
            Paragraph("<b>$250.00</b>", td_right),
            Paragraph("1-on-1 walkthrough of React review dashboard, PDF preview, manual rescan", td_left),
        ],
        [
            Paragraph("<b>TOTAL ONBOARDING FEE</b>", td_bold),
            Paragraph("<b>5.5 hrs</b>", td_center),
            Paragraph("<b>$0.00</b>", td_right),
            Paragraph("<b>$330.00</b>", td_right),
            Paragraph("<b>$1,200.00</b>", td_bold),
            Paragraph("<b>Complete turnkey automation setup (One-time investment)</b>", td_bold),
        ]
    ]
    t_setup = Table(setup_pdf_data, colWidths=[120, 45, 55, 55, 65, 164])
    t_setup.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_setup)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 2: RECURRING INFRASTRUCTURE COSTS ──
    elements.append(Paragraph("2. ITEMIZED MONTHLY PRODUCTION OPERATING COSTS (8GB / 200GB SSD)", h1_style))
    
    infra_pdf_data = [
        [
            Paragraph("Component / Service", th_style),
            Paragraph("Specification & Role", th_style),
            Paragraph("DigitalOcean", th_style),
            Paragraph("GCP (Google Cloud)", th_style),
        ],
        [
            Paragraph("<b>Cloud Server (VPS)</b>", td_left),
            Paragraph("4 vCPU, 8 GB RAM, 160 GB NVMe SSD (5TB Transfer)", td_left),
            Paragraph("$48.00 / mo", td_right),
            Paragraph("$68.50 / mo", td_right),
        ],
        [
            Paragraph("<b>NVMe Storage Volume</b>", td_left),
            Paragraph("+40 GB NVMe Block Storage (200GB Total SSD)", td_left),
            Paragraph("$4.00 / mo", td_right),
            Paragraph("$8.00 / mo", td_right),
        ],
        [
            Paragraph("<b>Dedicated Static IPv4</b>", td_left),
            Paragraph("1 Public IP for Xero Webhooks & SSL Server", td_left),
            Paragraph("$0.00 (Included)", td_right),
            Paragraph("$3.65 / mo", td_right),
        ],
        [
            Paragraph("<b>Gemini AI Token Parsing</b>", td_left),
            Paragraph("5,000 invoices/mo (~5M tokens @ $0.075/1M)", td_left),
            Paragraph("$0.75 / mo", td_right),
            Paragraph("$0.75 / mo", td_right),
        ],
        [
            Paragraph("<b>Gemini API Registration Fee</b>", td_left),
            Paragraph("Google AI Studio account setup", td_left),
            Paragraph("$0.00 (Free)", td_right),
            Paragraph("$0.00 (Free)", td_right),
        ],
        [
            Paragraph("<b>Xero Platform API</b>", td_left),
            Paragraph("Multi-tenant Custom Connection & OAuth2", td_left),
            Paragraph("$0.00 (Free)", td_right),
            Paragraph("$0.00 (Free)", td_right),
        ],
        [
            Paragraph("<b>Email Intake (IMAP)</b>", td_left),
            Paragraph("Direct IMAP IDLE connection to client's mailbox", td_left),
            Paragraph("$0.00 (Free)", td_right),
            Paragraph("$0.00 (Free)", td_right),
        ],
        [
            Paragraph("<b>Domain & SSL Security</b>", td_left),
            Paragraph("Custom domain (.com) + Let's Encrypt SSL", td_left),
            Paragraph("$1.25 / mo", td_right),
            Paragraph("$1.25 / mo", td_right),
        ],
        [
            Paragraph("<b>Automated Backups</b>", td_left),
            Paragraph("Daily SQLite DB snapshots to Cloudflare R2 / S3", td_left),
            Paragraph("$2.00 / mo", td_right),
            Paragraph("$5.20 / mo", td_right),
        ],
        [
            Paragraph("<b>TOTAL REAL MONTHLY COST</b>", td_bold),
            Paragraph("<b>Complete 24/7 multi-company hosting cluster</b>", td_bold),
            Paragraph("<b>$56.00 / month</b>", td_bold),
            Paragraph("<b>$87.35 / month</b>", td_bold),
        ]
    ]
    t_infra = Table(infra_pdf_data, colWidths=[130, 204, 85, 85])
    t_infra.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#DBEAFE")),
        ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
    ]))
    elements.append(t_infra)
    elements.append(Spacer(1, 6))
    
    elements.append(PageBreak())
    
    # ── SECTION 3: COMMERCIAL PACKAGES & PROFIT MODEL ──
    elements.append(Paragraph("3. RECOMMENDED CLIENT PACKAGES & SAAS MARGINS", h1_style))
    elements.append(Paragraph(
        "Commercial packages structuring setup fees and recurring monthly retainers per connected company:",
        body_style
    ))
    
    pkg_pdf_data = [
        [
            Paragraph("Package Tier", th_style),
            Paragraph("Setup Fee", th_style),
            Paragraph("Monthly Retainer", th_style),
            Paragraph("Included Volume", th_style),
            Paragraph("Net Margin", th_style),
            Paragraph("Ideal Client", th_style),
        ],
        [
            Paragraph("<b>Standard Business</b><br/><i>(Recommended)</i>", td_left),
            Paragraph("$1,200.00", td_right),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("Up to 500 inv/mo", td_center),
            Paragraph("<b>$185.10 / mo (93%)</b>", td_right),
            Paragraph("Single SME with standard monthly bills", td_left),
        ],
        [
            Paragraph("<b>Growth / High-Vol</b>", td_left),
            Paragraph("$1,800.00", td_right),
            Paragraph("$349.00 / mo", td_right),
            Paragraph("Up to 1,500 inv/mo", td_center),
            Paragraph("<b>$333.50 / mo (95%)</b>", td_right),
            Paragraph("Trading / E-commerce with heavy invoice volume", td_left),
        ],
        [
            Paragraph("<b>Accounting Agency</b>", td_left),
            Paragraph("$2,800.00", td_right),
            Paragraph("$699.00 / mo", td_right),
            Paragraph("Up to 3,500 inv/mo", td_center),
            Paragraph("<b>$675.00 / mo (96%)</b>", td_right),
            Paragraph("Accounting firm managing up to 10 entities", td_left),
        ]
    ]
    t_pkg = Table(pkg_pdf_data, colWidths=[110, 65, 75, 75, 84, 95])
    t_pkg.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (4,1), (4,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_pkg)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 4: CLIENT ROI MODEL ──
    elements.append(Paragraph("4. CLIENT RETURN ON INVESTMENT (ROI) JUSTIFICATION", h1_style))
    elements.append(Paragraph(
        "Why the client will happily pay the $1,200 setup fee + $199/month retainer:",
        body_style
    ))
    
    roi_pdf_data = [
        [
            Paragraph("Parameter", th_style),
            Paragraph("Manual Data Entry", th_style),
            Paragraph("Automated System", th_style),
            Paragraph("Net Client Benefit", th_style),
        ],
        [
            Paragraph("<b>Monthly Invoice Volume</b>", td_left),
            Paragraph("500 invoices", td_center),
            Paragraph("500 invoices", td_center),
            Paragraph("100% automated intake", td_center),
        ],
        [
            Paragraph("<b>Time per Invoice</b>", td_left),
            Paragraph("5.0 minutes / bill", td_center),
            Paragraph("0.5 minutes (30s review)", td_center),
            Paragraph("<b>4.5 min saved / bill (90%)</b>", td_center),
        ],
        [
            Paragraph("<b>Monthly Human Labor Hours</b>", td_left),
            Paragraph("41.67 Hours", td_center),
            Paragraph("4.17 Hours", td_center),
            Paragraph("<b>37.5 Hours freed up / mo</b>", td_center),
        ],
        [
            Paragraph("<b>Monthly Labor Cost (@ $25/hr)</b>", td_left),
            Paragraph("$1,041.67", td_right),
            Paragraph("$104.17", td_right),
            Paragraph("<b>$937.50 / month saved</b>", td_right),
        ],
        [
            Paragraph("<b>Monthly Automation Retainer</b>", td_left),
            Paragraph("$0.00", td_right),
            Paragraph("$199.00", td_right),
            Paragraph("-$199.00 / month fee", td_right),
        ],
        [
            Paragraph("<b>NET MONTHLY CLIENT BENEFIT</b>", td_bold),
            Paragraph("$1,041.67", td_bold),
            Paragraph("$303.17", td_bold),
            Paragraph("<b>+$738.50 / MONTH SAVINGS</b>", td_bold),
        ],
        [
            Paragraph("<b>ANNUAL NET VALUE TO CLIENT</b>", td_bold),
            Paragraph("$12,500.00 / yr", td_bold),
            Paragraph("$3,638.00 / yr", td_bold),
            Paragraph("<b>+$8,862.00 / YEAR NET GAIN</b>", td_bold),
        ]
    ]
    t_roi = Table(roi_pdf_data, colWidths=[140, 110, 110, 144])
    t_roi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-3), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_roi)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 5: EXECUTIVE SUMMARY ──
    elements.append(Paragraph("5. EXECUTIVE SUMMARY FOR CLIENT QUOTATION", h1_style))
    elements.append(Paragraph(
        "• <b>One-Time Setup Fee ($1,200):</b> Covers Xero OAuth app creation, IMAP mailbox configuration, Gemini AI key rotation, 20-invoice pilot testing, and staff training.<br/>"
        "• <b>Monthly Managed Retainer ($199/mo):</b> Covers 24/7 cloud server hosting, automated SQLite backups, Gemini AI parsing tokens, Xero sync maintenance, and ongoing support.<br/>"
        "• <b>Client Payback Period:</b> Under <b>1.5 months</b>. The client saves <b>$738.50 every month ($8,862/year)</b> while delivering you a <b>93% recurring profit margin</b>.",
        body_style
    ))
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    print(f"PDF report successfully created at: {out_path}")


# ==============================================================================
# 3. MARKDOWN REPORT GENERATOR (With Setup Labor, IMAP & Gemini Registration)
# ==============================================================================
def create_md_report(out_path):
    md_content = """# Xero Multi-Tenant Invoice Automation — Cost Estimation & Client Setup Proposal

**Document Purpose:** Commercial Feasibility, Onboarding Labor Breakdown, API Registration & Client Fee Proposal  
**Target Specification:** 8 GB RAM | 160–200 GB NVMe SSD | 4 vCPU  
**Target Architecture:** Multi-Tenant Node.js + SQLite WAL + React SPA + Bull/Redis + Gemini AI  

---

## 1. Itemized Onboarding & Setup Costs (Helping Clients Set Up)

When onboarding each new company/client, here is the exact labor and registration breakdown:

| Onboarding & Setup Activity | Est. Time | Direct Vendor Fee | Internal Labor Value (@ $60/hr) | **Recommended Billable to Client** | Key Deliverables Included |
|---|---|---|---|---|---|
| **1. Xero OAuth2 & Custom Connection** | 1.5 hrs | **$0.00** | $90.00 | **$250.00** | Create Xero Dev App, generate Client ID/Secret, tenant ID mapping, Account Code 310 & GST tax rate mapping. |
| **2. IMAP Mailbox & Security Setup** | 1.0 hr | **$0.00** | $60.00 | **$200.00** | Google Workspace / M365 App Password generation, 2FA security compliance, 24/7 IMAP IDLE real-time testing. |
| **3. Gemini AI API & Key Registration** | 0.5 hrs | **$0.00** | $30.00 | **$150.00** | Google AI Studio account creation, API key generation, 15 RPM rate limiting, key rotation pool setup. |
| **4. Pilot Ingestion & Extraction Tuning** | 1.5 hrs | **$0.00** | $90.00 | **$350.00** | Ingesting 15–20 real vendor PDFs, tuning prompts for multi-line items, currencies, and PayNow/Bank details. |
| **5. Client Finance Staff Handover** | 1.0 hr | **$0.00** | $60.00 | **$250.00** | 1-on-1 live walkthrough of the React review dashboard, PDF preview iframe, approval queue, and manual rescan. |
| **TOTAL ONBOARDING INVESTMENT** | **5.5 hrs** | **$0.00** | **$330.00** | **$1,200.00 (One-Time Fee)** | **Turnkey Automated Pipeline Handover** |

> **Note on Vendor Registration Fees:**  
> • **Google Gemini API Registration:** **$0.00** (Google does not charge an account creation or registration fee).  
> • **Xero Developer Registration:** **$0.00** (Creating Custom Connections and OAuth2 apps is completely free).  
> • **IMAP Mailbox:** **$0.00** if using the client's existing inbox (or ~$6.00/mo if provisioning a new Google Workspace license).

---

## 2. Monthly Direct Operating Costs (8GB RAM / 160–200GB SSD)

| Infrastructure Component | Specification / Capacity | **DigitalOcean Route** | **GCP (Google Cloud)** |
|---|---|---|---|
| **Cloud Server (VPS)** | 4 vCPU / 8 GB RAM / 160 GB NVMe SSD | **$48.00 / mo** | $68.50 / mo (2 vCPU) |
| **NVMe Storage Volume** | +40 GB NVMe Block Storage (200GB Total SSD) | **$4.00 / mo** | $8.00 / mo |
| **Dedicated Static IPv4** | 1 Public IP for Xero Webhooks & SSL Server | **$0.00** (Included) | $3.65 / mo |
| **Bandwidth / Transfer** | 5,000 GB (5TB) Included on DigitalOcean | **$0.00** (Included) | ~$8.00 / mo |
| **Gemini AI Token Parsing** | 5,000 invoices/mo (~5M tokens @ $0.075/1M) | **$0.75 / mo** | $0.75 / mo |
| **Xero Platform API** | Multi-tenant Custom Connection & OAuth2 | **$0.00** (Free Developer) | $0.00 (Free Developer) |
| **Domain & SSL Security** | Custom domain (.com) + Let's Encrypt SSL | **$1.25 / mo** | $1.25 / mo |
| **Automated Backups** | Daily SQLite DB snapshots (Cloudflare R2 / S3) | **$2.00 / mo** | $5.20 / mo |
| **TOTAL DIRECT OPERATING EXPENSE** | **Complete 24/7 multi-company hosting cluster** | **~$56.00 / month** | **~$87.35 / month** |

---

## 3. Recommended Client Packages & SaaS Profit Margins

| Package Tier | One-Time Setup Fee | Monthly Retainer | Included Volume | Real Server Cost | Net Monthly Margin | Target Client Profile |
|---|---|---|---|---|---|---|
| **Standard Business (Recommended)** | **$1,200.00** | **$199.00 / mo** | Up to 500 inv/mo | ~$13.90 / mo | **$185.10 / mo (93%)** | Single SME with standard monthly bills |
| **Growth / High-Volume** | **$1,800.00** | **$349.00 / mo** | Up to 1,500 inv/mo | ~$15.50 / mo | **$333.50 / mo (95%)** | Trading, e-commerce, multi-entity |
| **Accounting Agency Tier** | **$2,800.00** | **$699.00 / mo** | Up to 3,500 inv/mo | ~$24.00 / mo | **$675.00 / mo (96%)** | Accounting firm managing multiple Xero books |

---

## 4. Multi-Company SaaS Profitability Scale

Because **one $56/mo server** easily hosts **20 to 50 connected companies**, your recurring profits grow exponentially:

| Active Companies | Monthly Retainer / Co. | Gross Monthly Revenue | Total Server & AI Cost | **Net Monthly Profit** | **Gross Profit Margin** |
|---|---|---|---|---|---|
| **5 Companies** | $199.00 / mo | **$995.00 / mo** | $56.75 / mo | **+$938.25 / month** | **94.3%** |
| **15 Companies** | $199.00 / mo | **$2,985.00 / mo** | $57.50 / mo | **+$2,927.50 / month** | **98.1%** |
| **35 Companies** | $199.00 / mo | **$6,965.00 / mo** | $62.00 / mo | **+$6,903.00 / month** | **99.1%** |
| **75 Companies** | $199.00 / mo | **$14,925.00 / mo** | $120.00 / mo (Scaled) | **+$14,805.00 / month** | **99.2%** |

---

## 5. Client ROI & Justification (Why the Client Will Buy)

1. **Manual Entry Cost:** 500 invoices × 5 minutes = **41.6 hours/month**. At $25/hr, manual data entry costs **$1,041.67 / month**.
2. **Automated Cost:** 30-second review on the dashboard = 4.1 hours/month ($104.17 labor) + $199 retainer = **$303.17 / month total**.
3. **Net Monthly Savings:** The client saves **$738.50 every month ($8,862 / year)**.
4. **Payback Period:** Under **1.5 months** (even including the $1,200 setup fee).
"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Markdown report successfully created at: {out_path}")


if __name__ == "__main__":
    # Generate in root /report
    create_excel_report(os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.xlsx"))
    create_pdf_report(os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.pdf"))
    create_md_report(os.path.join(REPORT_DIR, "COST_ESTIMATION_AND_PRICING_REPORT.md"))

    # Generate in /report/arch_recommendation
    create_excel_report(os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.xlsx"))
    create_pdf_report(os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.pdf"))
    create_md_report(os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.md"))

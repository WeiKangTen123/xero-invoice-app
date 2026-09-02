import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

TARGET_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report/arch_recommendation"
XLSX_PATH = os.path.join(TARGET_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.xlsx")
PDF_PATH = os.path.join(TARGET_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.pdf")
MD_PATH = os.path.join(TARGET_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.md")

# ==============================================================================
# 1. EXCEL REPORT GENERATOR
# ==============================================================================
def create_excel_report():
    wb = openpyxl.Workbook()
    
    NAVY_DARK = "1E3A8A"
    NAVY_LIGHT = "DBEAFE"
    SLATE_DARK = "334155"
    SLATE_LIGHT = "F8FAFC"
    GREEN_ACCENT = "0D9488"
    GREEN_LIGHT = "CCFBF1"
    GRAY_TEXT = "64748B"
    WHITE = "FFFFFF"
    
    font_title = Font(name="Calibri", size=16, bold=True, color=NAVY_DARK)
    font_subtitle = Font(name="Calibri", size=11, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=12, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=11, color="1E293B")
    font_data_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_total = Font(name="Calibri", size=11, bold=True, color=NAVY_DARK)
    
    fill_sec_hdr = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=SLATE_DARK, end_color=SLATE_DARK, fill_type="solid")
    fill_alt_row = PatternFill(start_color=SLATE_LIGHT, end_color=SLATE_LIGHT, fill_type="solid")
    fill_highlight = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    fill_total = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    total_border = Border(
        top=Side(style='thin', color='1E3A8A'),
        bottom=Side(style='double', color='1E3A8A'),
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    # -------------------------------------------------------------
    # TAB 1: ARCHITECTURE COMPARISON MATRIX
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Architecture Comparison"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "XERO AUTOMATION — SYSTEM ARCHITECTURE & SERVER DESIGN COMPARISON"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_left
    
    ws1.merge_cells("A2:F2")
    ws1["A2"] = "Technical & Financial Evaluation of Monolith vs. Decoupled 3-Tier vs. Cloud-Native Microservices"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = align_left
    
    # Metadata
    ws1["A4"] = "Current Target Spec:"
    ws1["B4"] = "8 GB RAM | 160-200 GB NVMe SSD | 4 vCPU"
    ws1["A5"] = "Evaluation Date:"
    ws1["B5"] = "August 2026"
    ws1["D4"] = "System Type:"
    ws1["E4"] = "Multi-Company Real-Time Email-to-Xero Sync"
    ws1["D5"] = "Base Currency:"
    ws1["E5"] = "USD ($)"
    for r in [4, 5]:
        ws1[f"A{r}"].font = font_data_bold
        ws1[f"B{r}"].font = font_data
        ws1[f"D{r}"].font = font_data_bold
        ws1[f"E{r}"].font = font_data

    ws1.merge_cells("A7:F7")
    ws1["A7"] = "ARCHITECTURE OPTIONS COMPARISON MATRIX"
    ws1["A7"].font = font_sec_hdr
    ws1["A7"].fill = fill_sec_hdr

    headers_arch = ["Architecture Design", "Monthly Infra Cost", "Capacity (Companies)", "Crash Isolation / Reliability", "Database & PDF Setup", "Strategic Verdict"]
    for c_idx, h in enumerate(headers_arch, 1):
        cell = ws1.cell(row=8, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    arch_rows = [
        ["Design 1: All-in-One Monolith (Current Setup)", 56.00, "1 to 30 Companies", "Low (Node crash halts UI/DB)", "Embedded SQLite WAL + Local SSD PDFs", "BEST FOR PHASE 1 (Maximum profit & fast launch)"],
        ["Design 2: Decoupled 3-Tier Production SaaS", 135.00, "30 to 150 Companies", "High (Separate Managed DB & Queue)", "Managed Postgres + Cloudflare R2 PDFs + Redis", "RECOMMENDED AT SCALE (Zero data loss risk)"],
        ["Design 3: Cloud-Native Microservices", 450.00, "200+ Enterprise Orgs", "Maximum (Autoscaling K8s / Cloud Run)", "AWS Aurora Postgres + S3 Glacier + Kafka", "OVERKILL FOR CURRENT STAGE (Too complex)"],
    ]

    for r_idx, row_vals in enumerate(arch_rows, start=9):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx == 2:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            elif c_idx == 6 and "BEST" in str(val):
                cell.font = font_total
                cell.fill = fill_highlight
                cell.alignment = align_left
            else:
                cell.font = font_data
                cell.alignment = align_center if c_idx == 3 else align_left

    # Detailed Subsystems Table
    ws1.merge_cells("A14:F14")
    ws1["A14"] = "SUBSYSTEM IMPLEMENTATION DETAILS BY ARCHITECTURE"
    ws1["A14"].font = font_sec_hdr
    ws1["A14"].fill = fill_sec_hdr

    sub_headers = ["Subsystem Layer", "Design 1: Monolith (Current)", "Design 2: Decoupled 3-Tier", "Design 3: Cloud-Native", "Why It Matters in Real Life", "Tech Stack"]
    for c_idx, h in enumerate(sub_headers, 1):
        cell = ws1.cell(row=15, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    sub_data = [
        ["Frontend UI", "Served statically by Express port 4000", "Cloudflare Pages / Vercel Global Edge", "Cloudflare Edge + Global WAF", "Decoupling speeds up page load globally", "React SPA + Vite"],
        ["Backend API & Watcher", "1 Node.js process on VPS", "2 App Server Containers + Bull Workers", "Autoscaling Kubernetes / Cloud Run", "Handles 50+ persistent IMAP TLS connections", "Express + node-imap"],
        ["Database Layer", "SQLite WAL (data/app.db)", "Managed PostgreSQL (DigitalOcean)", "AWS Aurora Serverless / Cloud SQL", "Postgres ensures ACID safety across cluster", "better-sqlite3 -> pg"],
        ["Invoice PDF Storage", "Local SSD (data/users/{id}/pdfs/)", "Cloudflare R2 Object Storage", "AWS S3 Multi-Region Glacier", "Cloud storage prevents running out of disk", "Local FS -> S3/R2 API"],
        ["Job Queue & Worker", "Disk queue + in-memory Bull", "Dedicated Managed Redis Instance", "Redis Cluster + Kafka Event Bus", "Absorbs Monday morning email traffic surges", "ioredis + Bull Queue"],
        ["Disaster Recovery", "Daily script to S3/Cloud", "Automated Point-in-Time DB Recovery", "Multi-AZ active-active failover", "Protects financial tax data compliance", "sqlite3 .backup / pg_dump"]
    ]

    for r_idx, row_vals in enumerate(sub_data, start=16):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            cell.font = font_data_bold if c_idx == 1 else font_data
            cell.alignment = align_left

    # -------------------------------------------------------------
    # TAB 2: $100 - $150 PRODUCTION INFRASTRUCTURE BUDGET
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Production Infra ($100-$150)")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "REALISTIC PRODUCTION CLOUD BUDGET ($100 – $150 / MONTH REALITY)"
    ws2["A1"].font = font_title
    
    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Complete breakdown including Staging VM, Managed Redis, Snapshots, Load Balancer & High Availability"
    ws2["A2"].font = font_subtitle
    
    headers_budget = ["Infrastructure Component", "Technical Purpose / Specification", "DigitalOcean Cost", "GCP Cost", "Necessity Level", "Notes"]
    for c_idx, h in enumerate(headers_budget, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    budget_items = [
        ["Primary Production VPS", "4 vCPU / 8 GB RAM / 160 GB NVMe SSD", 48.00, 68.50, "Mandatory", "Runs 20-50 company IMAP watchers & dashboard"],
        ["Staging / UAT Test Server", "1 vCPU / 2 GB RAM / 50 GB SSD", 12.00, 22.00, "Recommended", "Test updates/Xero sync before touching live data"],
        ["Dedicated NVMe Storage", "+40 GB block storage (200GB Total)", 4.00, 8.00, "Mandatory", "Stores 400,000+ invoice PDFs + WAL database"],
        ["Managed Redis Queue", "Dedicated 1GB Memory Redis Instance", 15.00, 35.00, "Recommended", "Prevents memory starvation on background worker"],
        ["Automated Daily Backups", "1-Click VPS Snapshots + Disaster Recovery", 9.60, 12.00, "Mandatory", "Full server image recovery in < 5 minutes"],
        ["Cloud Load Balancer / SSL", "HTTPS termination + Health Check routing", 12.00, 18.00, "Optional/HA", "Zero-downtime rolling deploys"],
        ["Static Public IPv4", "Dedicated IP for Webhook endpoints", 0.00, 3.65, "Mandatory", "Free on DO, charged continuously on GCP"],
        ["Network Egress Bandwidth", "High-speed PDF streaming & sync data", 0.00, 10.00, "Mandatory", "5,000 GB (5TB) included on DO; Pay-per-GB on GCP"],
        ["Domain & DNS Hosting", "Custom Domain Name (.com / .app)", 1.25, 1.25, "Mandatory", "Amortized $15/year"],
        ["Offsite Cold S3 Storage", "500 GB Cloudflare R2 / S3 Archival", 7.50, 10.00, "Mandatory", "Long-term 7-year tax compliance archival"]
    ]

    for r_idx, row_vals in enumerate(budget_items, start=5):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx in [3, 4]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            elif c_idx == 5:
                cell.font = font_data_bold if val == "Mandatory" else font_data
                cell.alignment = align_center
            else:
                cell.font = font_data
                cell.alignment = align_left

    tot_row = len(budget_items) + 5
    ws2.cell(row=tot_row, column=1, value="TOTAL REALISTIC CLOUD INFRASTRUCTURE BUDGET").font = font_total
    ws2.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=2)
    for c in range(1, 3):
        ws2.cell(row=tot_row, column=c).fill = fill_total
        ws2.cell(row=tot_row, column=c).border = total_border

    c_do_tot = ws2.cell(row=tot_row, column=3, value=f"=SUM(C5:C{tot_row-1})")
    c_do_tot.font = font_total
    c_do_tot.fill = fill_total
    c_do_tot.border = total_border
    c_do_tot.alignment = align_right
    c_do_tot.number_format = "$#,##0.00"

    c_gcp_tot = ws2.cell(row=tot_row, column=4, value=f"=SUM(D5:D{tot_row-1})")
    c_gcp_tot.font = font_total
    c_gcp_tot.fill = fill_total
    c_gcp_tot.border = total_border
    c_gcp_tot.alignment = align_right
    c_gcp_tot.number_format = "$#,##0.00"

    for c in [5, 6]:
        ws2.cell(row=tot_row, column=c).fill = fill_total
        ws2.cell(row=tot_row, column=c).border = total_border

    # -------------------------------------------------------------
    # TAB 3: HIDDEN TECHNICAL COSTS & THIRD-PARTY SAAS
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Hidden Technical Costs")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:F1")
    ws3["A1"] = "TECHNICAL HIDDEN COSTS & OPERATIONAL EXPENSES AUDIT"
    ws3["A1"].font = font_title
    
    ws3.merge_cells("A2:F2")
    ws3["A2"] = "Every operational expense often overlooked in initial estimates"
    ws3["A2"].font = font_subtitle
    
    headers_hidden = ["Category", "Tool / Service Line Item", "Typical Monthly Cost", "Why It Is Necessary", "Trigger / Volume Condition", "Mitigation Strategy"]
    for c_idx, h in enumerate(headers_hidden, 1):
        cell = ws3.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    hidden_items = [
        ["OCR & Scanned Invoices", "Google Cloud Vision / Gemini Multimodal", 15.00, "Extracts data from image/scanned PDFs where pdf-parse yields 0 text", "When suppliers send phone photos/scans", "Fall back to Gemini Flash Vision ($0.002/scan)"],
        ["Transactional Email", "Resend / Postmark / SendGrid", 20.00, "Sends review alerts, invite links & daily summaries to clients", "When active users exceed free tier (100/day)", "Use Amazon SES ($0.10 per 1k emails) to cut cost"],
        ["Error Tracking (APM)", "Sentry.io (Team Plan)", 26.00, "Instant alerts on unhandled crashes, dead sockets, or failed syncs", "Essential when running 24/7 client SLAs", "Use Sentry Developer Free tier for <5k errors/mo"],
        ["Uptime & SMS Alerts", "BetterStack / Uptime Kuma", 15.00, "SMS/Phone call alerts to developer when server or IMAP drops", "Immediate incident response", "Free Uptime Kuma + Telegram bot alerts"],
        ["Security & WAF", "Cloudflare Pro", 20.00, "Web application firewall, rate-limiting & DDoS protection", "Protects financial portal from brute force", "Cloudflare Free tier is sufficient for early stage"],
        ["Secrets Management", "Doppler / AWS KMS", 18.00, "Secure management of encrypted Xero tokens and AES keys", "Multi-developer team sync", "Local encrypted .env for single dev"],
        ["Maintenance Labor", "Developer On-Call & Format Drift", 150.00, "Adjusting prompts for weird invoice formats & Xero re-auth", "2-4 hours / month ongoing support", "Bake maintenance retainer into client contract"]
    ]

    for r_idx, row_vals in enumerate(hidden_items, start=5):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx == 3:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            else:
                cell.font = font_data
                cell.alignment = align_left

    # -------------------------------------------------------------
    # TAB 4: SAAS PRICING & MULTI-TENANT PROFIT MODEL
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="SaaS Profit & Pricing Model")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("A1:G1")
    ws4["A1"] = "MULTI-COMPANY COMMERCIAL PRICING & PROFIT PROJECTIONS"
    ws4["A1"].font = font_title
    
    ws4.merge_cells("A2:G2")
    ws4["A2"] = "Financial returns across various scale milestones on recommended infrastructure"
    ws4["A2"].font = font_subtitle
    
    headers_scale = ["Scale Milestone", "Connected Companies", "Retainer Fee / Co.", "Gross Monthly Revenue", "Total Cloud Infra Cost", "Total Hidden & AI Costs", "Net Monthly Profit"]
    for c_idx, h in enumerate(headers_scale, 1):
        cell = ws4.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    scale_data = [
        ["Phase 1: Initial Launch", 5, 199.00, "=B5*C5", 56.00, 15.00, "=D5-E5-F5"],
        ["Phase 1: Expansion", 15, 199.00, "=B6*C6", 56.00, 25.00, "=D6-E6-F6"],
        ["Phase 2: Decoupled 3-Tier", 35, 199.00, "=B7*C7", 135.00, 45.00, "=D7-E7-F7"],
        ["Phase 2: High Scale", 75, 199.00, "=B8*C8", 145.00, 75.00, "=D8-E8-F8"],
        ["Phase 3: Enterprise Cluster", 150, 199.00, "=B9*C9", 250.00, 150.00, "=D9-E9-F9"],
    ]

    for r_idx, row_vals in enumerate(scale_data, start=5):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx in [3, 4, 5, 6, 7]:
                cell.font = font_total if c_idx == 7 else font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
                if c_idx == 7:
                    cell.fill = fill_highlight
            else:
                cell.font = font_data
                cell.alignment = align_center

    # Adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str += "    "
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
    wb.save(XLSX_PATH)
    print(f"Architecture Excel report created at: {XLSX_PATH}")


# ==============================================================================
# 2. PDF REPORT GENERATOR
# ==============================================================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Header (pages 2+)
        if self._pageNumber > 1:
            self.drawString(54, 11 * inch - 36, "XERO INVOICE AUTOMATION — SYSTEM ARCHITECTURE & COST REPORT")
            self.drawRightString(8.5 * inch - 54, 11 * inch - 36, "CONFIDENTIAL")
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(54, 11 * inch - 42, 8.5 * inch - 54, 11 * inch - 42)
            
        # Footer
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(54, 46, 8.5 * inch - 54, 46)
        
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawString(54, 32, "Technical Architecture & Commercial Pricing Model | August 2026")
        self.drawRightString(8.5 * inch - 54, 32, page_str)
        self.restoreState()


def create_pdf_report():
    doc = SimpleDocTemplate(
        PDF_PATH,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    c_navy = colors.HexColor("#1E3A8A")
    c_slate = colors.HexColor("#334155")
    c_body = colors.HexColor("#1E293B")
    c_teal = colors.HexColor("#0D9488")
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=17, leading=21,
        textColor=c_navy, spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'],
        fontName='Helvetica-Oblique', fontSize=9, leading=12,
        textColor=colors.HexColor("#64748B"), spaceAfter=10
    )
    h1_style = ParagraphStyle(
        'SecH1', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10.5, leading=14,
        textColor=c_navy, spaceBefore=9, spaceAfter=4, keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyDark', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, leading=11,
        textColor=c_body, spaceAfter=4
    )
    th_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
        textColor=colors.white, alignment=1
    )
    td_left = ParagraphStyle(
        'TableCellLeft', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_bold = ParagraphStyle(
        'TableCellBold', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_right = ParagraphStyle(
        'TableCellRight', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=2
    )
    td_center = ParagraphStyle(
        'TableCellCenter', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=1
    )
    
    elements = []
    
    # Header Block
    elements.append(Paragraph("SYSTEM ARCHITECTURE & PRODUCTION COST RECOMMENDATION", title_style))
    elements.append(Paragraph("Multi-Company Real-Time Architecture, $100–$150 Cloud Sizing & Hidden Technical Costs", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=c_navy, spaceAfter=8))
    
    # Metadata Table
    meta_table_data = [
        [Paragraph("<b>Target Stack:</b> Multi-Tenant Node.js + SQLite WAL + React + Bull", td_left), Paragraph("<b>Date:</b> August 2026", td_left)],
        [Paragraph("<b>Target Hardware:</b> 8GB RAM | 160–200GB NVMe SSD | 4 vCPU", td_left), Paragraph("<b>Production Cost:</b> $56/mo (Monolith) to $120/mo (Decoupled)", td_left)],
        [Paragraph("<b>Recommended Cloud:</b> DigitalOcean Droplet (Best price & zero egress)", td_left), Paragraph("<b>Client Retainer Target:</b> $199.00 / company / month", td_left)]
    ]
    t_meta = Table(meta_table_data, colWidths=[270, 234])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 6))
    
    # Section 1: Architecture Comparison
    elements.append(Paragraph("1. ARCHITECTURE OPTIONS COMPARISON (MONOLITH VS 3-TIER)", h1_style))
    elements.append(Paragraph(
        "A structured comparison of how the system can be hosted in production, balancing simplicity, fault tolerance, and budget:",
        body_style
    ))
    
    arch_pdf_data = [
        [
            Paragraph("Architecture Design", th_style),
            Paragraph("Monthly Cost", th_style),
            Paragraph("Company Capacity", th_style),
            Paragraph("Database & Storage Setup", th_style),
            Paragraph("Verdict / Best For", th_style),
        ],
        [
            Paragraph("<b>Design 1: Monolith (Current)</b>", td_left),
            Paragraph("$56.00 / mo", td_right),
            Paragraph("1 to 30 Companies", td_center),
            Paragraph("Embedded SQLite WAL + Local NVMe SSD", td_left),
            Paragraph("<b>Phase 1 MVP (Fast launch, max margin)</b>", td_left),
        ],
        [
            Paragraph("<b>Design 2: Decoupled 3-Tier</b>", td_left),
            Paragraph("$135.00 / mo", td_right),
            Paragraph("30 to 150 Companies", td_center),
            Paragraph("Managed Postgres + Cloudflare R2 + Redis", td_left),
            Paragraph("<b>Recommended at Scale (Zero data loss)</b>", td_left),
        ],
        [
            Paragraph("<b>Design 3: Cloud-Native K8s</b>", td_left),
            Paragraph("$450.00+ / mo", td_right),
            Paragraph("200+ Enterprises", td_center),
            Paragraph("AWS Aurora + S3 Glacier + Kafka", td_left),
            Paragraph("Overkill for current stage", td_left),
        ]
    ]
    t_arch = Table(arch_pdf_data, colWidths=[120, 70, 85, 120, 109])
    t_arch.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (4,1), (4,1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_arch)
    elements.append(Spacer(1, 8))
    
    # Section 2: Real $100–$150 Cloud Budget
    elements.append(Paragraph("2. REALISTIC PRODUCTION CLOUD BUDGET ($100 – $150 / MO REALITY)", h1_style))
    
    budget_pdf_data = [
        [
            Paragraph("Component / Service", th_style),
            Paragraph("Technical Purpose", th_style),
            Paragraph("DigitalOcean", th_style),
            Paragraph("GCP (Google Cloud)", th_style),
            Paragraph("Necessity", th_style),
        ],
        [
            Paragraph("<b>Primary Production VPS</b>", td_left),
            Paragraph("4 vCPU / 8 GB RAM / 160 GB NVMe SSD", td_left),
            Paragraph("$48.00 / mo", td_right),
            Paragraph("$68.50 / mo", td_right),
            Paragraph("Mandatory", td_center),
        ],
        [
            Paragraph("<b>Staging / UAT Server</b>", td_left),
            Paragraph("1 vCPU / 2 GB RAM (Safe test environment)", td_left),
            Paragraph("$12.00 / mo", td_right),
            Paragraph("$22.00 / mo", td_right),
            Paragraph("Recommended", td_center),
        ],
        [
            Paragraph("<b>Dedicated NVMe Volume</b>", td_left),
            Paragraph("+40 GB NVMe Storage (200GB Total)", td_left),
            Paragraph("$4.00 / mo", td_right),
            Paragraph("$8.00 / mo", td_right),
            Paragraph("Mandatory", td_center),
        ],
        [
            Paragraph("<b>Managed Redis Queue</b>", td_left),
            Paragraph("Dedicated memory queue for Bull worker", td_left),
            Paragraph("$15.00 / mo", td_right),
            Paragraph("$35.00 / mo", td_right),
            Paragraph("Recommended", td_center),
        ],
        [
            Paragraph("<b>Daily Automated Snapshots</b>", td_left),
            Paragraph("1-click full VPS image disaster recovery", td_left),
            Paragraph("$9.60 / mo", td_right),
            Paragraph("$12.00 / mo", td_right),
            Paragraph("Mandatory", td_center),
        ],
        [
            Paragraph("<b>Cloud Load Balancer / SSL</b>", td_left),
            Paragraph("Zero-downtime rolling reboot routing", td_left),
            Paragraph("$12.00 / mo", td_right),
            Paragraph("$18.00 / mo", td_right),
            Paragraph("Optional/HA", td_center),
        ],
        [
            Paragraph("<b>Static IPv4 & Bandwidth</b>", td_left),
            Paragraph("Dedicated IP + 5TB transfer on DO", td_left),
            Paragraph("$0.00 (Free)", td_right),
            Paragraph("~$13.65 / mo", td_right),
            Paragraph("Mandatory", td_center),
        ],
        [
            Paragraph("<b>Cold Archival Storage (R2/S3)</b>", td_left),
            Paragraph("500 GB 7-year tax compliance storage", td_left),
            Paragraph("$7.50 / mo", td_right),
            Paragraph("$10.00 / mo", td_right),
            Paragraph("Mandatory", td_center),
        ],
        [
            Paragraph("<b>TOTAL CLOUD BUDGET</b>", td_bold),
            Paragraph("<b>Complete High-Availability Cluster</b>", td_bold),
            Paragraph("<b>$108.10 / mo</b>", td_bold),
            Paragraph("<b>$187.15 / mo</b>", td_bold),
            Paragraph("<b>Production</b>", td_center),
        ]
    ]
    t_budget = Table(budget_pdf_data, colWidths=[110, 160, 75, 85, 74])
    t_budget.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#DBEAFE")),
        ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
    ]))
    elements.append(t_budget)
    elements.append(Spacer(1, 6))
    
    elements.append(PageBreak())
    
    # Section 3: Hidden Costs & Third-Party Services
    elements.append(Paragraph("3. HIDDEN TECHNICAL COSTS & OPERATIONAL REALITIES", h1_style))
    elements.append(Paragraph(
        "Critical technical services and operational realities required to run a commercial multi-client automation platform:",
        body_style
    ))
    
    hidden_pdf_data = [
        [
            Paragraph("Cost Item", th_style),
            Paragraph("Estimated Cost", th_style),
            Paragraph("Why It's Needed (Real-Life Failure Without It)", th_style),
            Paragraph("Mitigation Strategy", th_style),
        ],
        [
            Paragraph("<b>OCR / Scanned Image Invoices</b>", td_left),
            Paragraph("$15.00 / mo", td_right),
            Paragraph("Suppliers who send phone photos/scans yield 0 text in regular PDF parsers.", td_left),
            Paragraph("Use Gemini Flash Vision ($0.0025/scan).", td_left),
        ],
        [
            Paragraph("<b>Transactional Email (Alerts)</b>", td_left),
            Paragraph("$20.00 / mo", td_right),
            Paragraph("Sends review notices, user invitations, and sync reports to accounting staff.", td_left),
            Paragraph("Use Amazon SES ($0.10/1k emails).", td_left),
        ],
        [
            Paragraph("<b>Error Tracking (Sentry APM)</b>", td_left),
            Paragraph("$26.00 / mo", td_right),
            Paragraph("Catches dropped IMAP sockets, Xero 429 errors, and crashes before clients notice.", td_left),
            Paragraph("Free tier handles < 5,000 errors/mo.", td_left),
        ],
        [
            Paragraph("<b>Security & Cloudflare WAF</b>", td_left),
            Paragraph("$20.00 / mo", td_right),
            Paragraph("Protects the accounting review dashboard from brute-force login attacks.", td_left),
            Paragraph("Cloudflare Free tier is fine for early launch.", td_left),
        ],
        [
            Paragraph("<b>Xero Maintenance & Format Drift</b>", td_left),
            Paragraph("2-4 hrs / mo", td_center),
            Paragraph("Adjusting prompts for unusual invoice formats & refreshing expired OAuth tokens.", td_left),
            Paragraph("Covered by monthly client retainer fee.", td_left),
        ]
    ]
    t_hidden = Table(hidden_pdf_data, colWidths=[120, 65, 204, 115])
    t_hidden.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_hidden)
    elements.append(Spacer(1, 8))
    
    # Section 4: SaaS Profit Projections
    elements.append(Paragraph("4. MULTI-COMPANY SAAS PROFITABILITY MODEL", h1_style))
    elements.append(Paragraph(
        "Financial model illustrating gross revenue vs. all hard infrastructure and third-party SaaS tooling costs:",
        body_style
    ))
    
    scale_pdf_data = [
        [
            Paragraph("Milestone Stage", th_style),
            Paragraph("Companies", th_style),
            Paragraph("Monthly Retainer", th_style),
            Paragraph("Gross Revenue", th_style),
            Paragraph("Total Tech Costs", th_style),
            Paragraph("Net Monthly Profit", th_style),
            Paragraph("Gross Margin", th_style),
        ],
        [
            Paragraph("<b>Phase 1: Initial Pool</b>", td_left),
            Paragraph("5 Co.", td_center),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$995.00 / mo", td_right),
            Paragraph("$71.00 / mo", td_right),
            Paragraph("<b>$924.00 / mo</b>", td_right),
            Paragraph("<b>92.8%</b>", td_center),
        ],
        [
            Paragraph("<b>Phase 1: Expansion</b>", td_left),
            Paragraph("15 Co.", td_center),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$2,985.00 / mo", td_right),
            Paragraph("$81.00 / mo", td_right),
            Paragraph("<b>$2,904.00 / mo</b>", td_right),
            Paragraph("<b>97.3%</b>", td_center),
        ],
        [
            Paragraph("<b>Phase 2: Decoupled 3-Tier</b>", td_left),
            Paragraph("35 Co.", td_center),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$6,965.00 / mo", td_right),
            Paragraph("$180.00 / mo", td_right),
            Paragraph("<b>$6,785.00 / mo</b>", td_right),
            Paragraph("<b>97.4%</b>", td_center),
        ],
        [
            Paragraph("<b>Phase 2: Scaled Multi-Tenant</b>", td_left),
            Paragraph("75 Co.", td_center),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$14,925.00 / mo", td_right),
            Paragraph("$220.00 / mo", td_right),
            Paragraph("<b>$14,705.00 / mo</b>", td_right),
            Paragraph("<b>98.5%</b>", td_center),
        ]
    ]
    t_scale = Table(scale_pdf_data, colWidths=[110, 45, 65, 75, 65, 80, 64])
    t_scale.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (5,1), (5,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_scale)
    elements.append(Spacer(1, 8))
    
    # Section 5: Evolutionary Recommendation
    elements.append(Paragraph("5. STRATEGIC IMPLEMENTATION ROADMAP", h1_style))
    elements.append(Paragraph(
        "• <b>Start with Phase 1 Monolith:</b> Run the current All-in-One system on an 8GB DigitalOcean Droplet ($56/mo). It requires zero architectural rewrites and yields <b>>93% profit margin</b> immediately.<br/>"
        "• <b>Graduate to Phase 2 (3-Tier) at 25+ Companies:</b> When monthly revenue crosses $5,000/mo, connect Cloudflare R2 for unlimited PDF storage and switch SQLite to Managed PostgreSQL ($15/mo) for enterprise failover.",
        body_style
    ))
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    print(f"Architecture PDF report created at: {PDF_PATH}")


# ==============================================================================
# 3. MARKDOWN REPORT GENERATOR
# ==============================================================================
def create_md_report():
    md_content = """# Xero Invoice Automation — System Architecture & Production Cost Recommendation

**Document Purpose:** Technical Architecture Audit, Sizing Analysis ($100–$150 Cloud Budget) & Multi-Company Commercial Proposal  
**Target Specifications:** 8 GB RAM | 160–200 GB NVMe SSD | 4 vCPU  
**Generated Files in this folder:**
- **Excel Spreadsheet:** [`SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.xlsx`](file:///Users/weikangten/Desktop/xero-invoice-app-master/report/arch_recommendation/SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.xlsx)
- **Formal PDF Report:** [`SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.pdf`](file:///Users/weikangten/Desktop/xero-invoice-app-master/report/arch_recommendation/SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.pdf)

---

## 1. System Architecture Overview & Current Design

In your current codebase, the application operates as an **All-in-One Monolith**:

```
┌────────────────────────────────────────────────────────────────────────────────────────┐
│                   CURRENT ALL-IN-ONE MONOLITHIC ARCHITECTURE                           │
│                               Single 8GB VPS Server                                    │
├────────────────────────────────────────────────────────────────────────────────────────┤
│                                                                                        │
│   1. FRONTEND (React SPA in /ui):                                                      │
│      • Compiled to static HTML/JS/CSS (`ui/dist`).                                     │
│      • Served directly by Express on port 4000 via `app.use(express.static(UI_DIST))`. │
│                                                                                        │
│   2. BACKEND API & EMAIL WORKER (Node.js in /main):                                    │
│      • Express REST API (`/api/invoices`, `/api/dashboard`, `/api/setup`).              │
│      • IMAP Watcher Registry: 24/7 persistent TLS sockets to client email inboxes.    │
│      • Queue & Parser Worker: Disk-backed FIFO queue + Bull/Redis worker.              │
│                                                                                        │
│   3. EMBEDDED DATABASE (SQLite in /main/data/app.db):                                  │
│      • Direct disk-based SQLite with Write-Ahead Logging (WAL mode).                   │
│                                                                                        │
│   4. FILE STORAGE (Local Disk in /main/data/users/{userId}/pdfs/):                     │
│      • Stores raw supplier invoice PDFs on local disk.                                 │
│                                                                                        │
└────────────────────────────────────────────────────────────────────────────────────────┘
```

---

## 2. How Other Production SaaS Systems Are Designed (3 Industry Models)

| Architecture Model | Monthly Infra Cost | Target Company Capacity | Crash Isolation & Reliability | Database & PDF Storage | Recommended Verdict |
|---|---|---|---|---|---|
| **Design 1: All-in-One Monolith (Current Setup)** | **$56.00 / mo** | **1 to 30 Companies** | Low (Node crash halts UI & DB) | Embedded SQLite WAL + Local NVMe SSD | **BEST FOR PHASE 1 (Fast launch & max profit)** |
| **Design 2: Decoupled 3-Tier Production SaaS** | **$135.00 / mo** | **30 to 150 Companies** | High (Independent DB & queue) | Managed Postgres + Cloudflare R2 + Managed Redis | **RECOMMENDED AT SCALE (Zero data loss)** |
| **Design 3: Cloud-Native Microservices** | **$450.00+ / mo** | **200+ Enterprise Orgs** | Maximum (Autoscaling K8s) | AWS Aurora Postgres + S3 Glacier + Kafka | **OVERKILL for current stage** |

---

## 3. Realistic Production Cloud Budget ($100 – $150 / Month Reality)

When configuring a high-availability production cloud environment with staging, backups, and load balancing:

| Cloud Component | Technical Role | DigitalOcean Cost | GCP Cost | Necessity |
|---|---|---|---|---|
| **Primary Production VPS (8GB / 4 vCPU)** | 20–50 active IMAP sockets, API & worker | **$48.00 / mo** | $68.50 / mo | Mandatory |
| **Staging / UAT Test Server (2GB RAM)** | Safe testing of Xero updates & patches | **$12.00 / mo** | $22.00 / mo | Recommended |
| **Dedicated NVMe Storage (+40GB Volume)** | Reaches 200GB SSD for 400,000+ PDFs | **$4.00 / mo** | $8.00 / mo | Mandatory |
| **Managed Redis Instance** | Dedicated memory queue for Bull workers | **$15.00 / mo** | $35.00 / mo | Recommended |
| **Automated Daily Snapshots** | 1-Click full server image disaster recovery | **$9.60 / mo** | $12.00 / mo | Mandatory |
| **Cloud Load Balancer / SSL** | Zero-downtime rolling reboot routing | **$12.00 / mo** | $18.00 / mo | Optional/HA |
| **Static IPv4 Address** | Dedicated public IP for Xero webhooks | **$0.00** (Included) | $3.65 / mo | Mandatory |
| **Network Egress Bandwidth** | 5,000 GB (5TB) transfer on DigitalOcean | **$0.00** (Included) | ~$10.00 / mo | Mandatory |
| **Cold Archival Storage (R2 / S3)** | 500 GB 7-year legal tax compliance storage | **$7.50 / mo** | $10.00 / mo | Mandatory |
| **Domain & DNS Hosting** | Custom Domain Name (.com / .app) | **$1.25 / mo** | $1.25 / mo | Mandatory |
| **TOTAL CLOUD INFRASTRUCTURE BUDGET** | **Complete High-Availability Cluster** | **~$109.35 / mo** | **~$188.90 / mo** | |

---

## 4. Hidden Technical Costs & Operational Expenses

| Hidden Expense | Monthly Cost | Why It Is Necessary (Real-Life Failure Without It) | Mitigation Strategy |
|---|---|---|---|
| **OCR / Scanned Invoices** | **$15.00 / mo** | Suppliers who send phone photos or flat image scans yield 0 text in regular PDF parsers. | Fall back to Gemini Multimodal Flash Vision ($0.0025/scan). |
| **Transactional Email Alerts** | **$20.00 / mo** | Sends review alerts, user invitations, and daily summaries to accounting staff. | Use Amazon SES ($0.10 / 1k emails) to keep costs minimal. |
| **Error Tracking (Sentry APM)** | **$26.00 / mo** | Instant developer alerts when IMAP sockets drop or Xero rate limits occur. | Sentry Free tier covers < 5,000 errors/month. |
| **Uptime & SMS Alerts** | **$15.00 / mo** | SMS/Phone alerts to developer if server or IMAP crashes. | Use Uptime Kuma + Telegram bot alerts ($0). |
| **Security & Cloudflare WAF** | **$20.00 / mo** | Protects the accounting dashboard from brute-force login attacks. | Cloudflare Free tier is fine for early launch. |
| **Developer Maintenance** | **2–4 hrs / mo** | Adjusting prompts for weird invoice formats & Xero re-authorizations. | Factored into monthly client retainer. |

---

## 5. Multi-Company SaaS Profitability Projections

| Milestone Stage | Active Companies | Monthly Retainer / Co. | Gross Monthly Revenue | Total Cloud & Tooling Costs | Net Monthly Profit | Gross Profit Margin |
|---|---|---|---|---|---|---|
| **Phase 1: Initial Launch** | **5 Companies** | $199.00 / mo | **$995.00 / mo** | $71.00 / mo | **+$924.00 / mo** | **92.8%** |
| **Phase 1: Expansion** | **15 Companies** | $199.00 / mo | **$2,985.00 / mo** | $81.00 / mo | **+$2,904.00 / mo** | **97.3%** |
| **Phase 2: Decoupled 3-Tier** | **35 Companies** | $199.00 / mo | **$6,965.00 / mo** | $180.00 / mo | **+$6,785.00 / mo** | **97.4%** |
| **Phase 2: Scaled SaaS** | **75 Companies** | $199.00 / mo | **$14,925.00 / mo** | $220.00 / mo | **+$14,705.00 / mo** | **98.5%** |

---

## 6. Strategic Migration & Implementation Roadmap

```
┌────────────────────────────────────────────────────────────────────────────────────────┐
│ PHASE 1: LAUNCH (1 to 25 Companies) — KEEP CURRENT MONOLITH                            │
├────────────────────────────────────────────────────────────────────────────────────────┤
│ • Deploy current All-in-One architecture on an 8GB DigitalOcean Droplet ($56/mo).      │
│ • Keep SQLite with daily automated backups (`main/db/backup.js`) synced offsite.       │
│ • Zero code rewrites required; captures immediate >93% profit margin.                  │
└────────────────────────────────────────────────────────────────────────────────────────┘
                                      │
                                      ▼
┌────────────────────────────────────────────────────────────────────────────────────────┐
│ PHASE 2: GROWTH (25 to 100+ Companies) — DECOUPLE TO 3-TIER PRODUCTION SAAS           │
├────────────────────────────────────────────────────────────────────────────────────────┤
│ • Offload PDF storage to Cloudflare R2 ($5/mo for 250GB, zero egress fee).             │
│ • Migrate SQLite to Managed PostgreSQL ($15/mo on DO, `pg` driver is already in repo). │
│ • Add Managed Redis ($15/mo) for dedicated memory queue isolation.                    │
│ • Total Infrastructure Cost: ~$120 – $140/mo with enterprise 99.9% uptime reliability. │
└────────────────────────────────────────────────────────────────────────────────────────┘
```
"""
    with open(MD_PATH, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Architecture Markdown report created at: {MD_PATH}")


if __name__ == "__main__":
    create_excel_report()
    create_pdf_report()
    create_md_report()

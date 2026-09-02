import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

REPORT_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report"
XLSX_PATH = os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.xlsx")
PDF_PATH = os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.pdf")
MD_PATH = os.path.join(REPORT_DIR, "COST_ESTIMATION_AND_PRICING_REPORT.md")

# ==============================================================================
# 1. EXCEL REPORT GENERATOR (Multi-Company SaaS & 8GB / 100-200GB Scale)
# ==============================================================================
def create_excel_report():
    wb = openpyxl.Workbook()
    
    NAVY_DARK = "1E3A8A"
    NAVY_LIGHT = "DBEAFE"
    SLATE_DARK = "334155"
    SLATE_LIGHT = "F8FAFC"
    GREEN_ACCENT = "0D9488"
    GREEN_LIGHT = "CCFBF1"
    GRAY_TEXT = "64748B"
    WHITE = "FFFFFF"
    
    font_title = Font(name="Calibri", size=16, bold=True, color=NAVY_DARK)
    font_subtitle = Font(name="Calibri", size=11, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=12, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=11, color="1E293B")
    font_data_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_total = Font(name="Calibri", size=11, bold=True, color=NAVY_DARK)
    
    fill_sec_hdr = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=SLATE_DARK, end_color=SLATE_DARK, fill_type="solid")
    fill_alt_row = PatternFill(start_color=SLATE_LIGHT, end_color=SLATE_LIGHT, fill_type="solid")
    fill_highlight = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    fill_total = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    total_border = Border(
        top=Side(style='thin', color='1E3A8A'),
        bottom=Side(style='double', color='1E3A8A'),
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    # -------------------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY & MULTI-COMPANY SAAS PRICING
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary & Pricing"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "XERO INVOICE AUTOMATION — MULTI-TENANT COST & PRICING REPORT"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_left
    
    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Production System Analysis (8GB RAM / 100-200GB Storage), Real-Time Multi-Company Architecture"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = align_left
    
    # Metadata Block
    ws1["A4"] = "Current Target Spec:"
    ws1["B4"] = "8 GB RAM | 160-200 GB NVMe SSD | 4 vCPU"
    ws1["A5"] = "Target Currency:"
    ws1["B5"] = "USD ($)"
    ws1["D4"] = "Architecture:"
    ws1["E4"] = "Multi-Tenant Real-Time IMAP Watchers + Bull Worker + Xero OAuth"
    ws1["D5"] = "Target Capacity:"
    ws1["E5"] = "10 to 50 Connected Companies (Scalable to 100+)"
    for r in [4, 5]:
        ws1[f"A{r}"].font = font_data_bold
        ws1[f"B{r}"].font = font_data
        ws1[f"D{r}"].font = font_data_bold
        ws1[f"E{r}"].font = font_data

    # Section 1: Multi-Company Commercial Pricing Packages
    ws1.merge_cells("A7:G7")
    ws1["A7"] = "CLIENT CHARGING MODELS (HOW TO PRICE PER CONNECTED COMPANY)"
    ws1["A7"].font = font_sec_hdr
    ws1["A7"].fill = fill_sec_hdr
    ws1["A7"].alignment = align_left

    headers_pricing = [
        "Package Tier", "One-Time Setup Fee", "Monthly Retainer / Co.", 
        "Monthly Invoice Limit", "Storage & Features", "Est. Value Delivered", "Target Client Profile"
    ]
    for col_num, h in enumerate(headers_pricing, 1):
        cell = ws1.cell(row=8, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    pricing_data = [
        ["Standard Business Tier (Recommended)", 1200.00, 199.00, "Up to 500 invoices/mo", "Live Dashboard + IMAP Watcher + 10GB PDF Storage", "$1,040 / mo labor saved", "Single company with 50-500 monthly bills"],
        ["Growth / High-Volume Tier", 1500.00, 349.00, "Up to 1,500 invoices/mo", "Multi-Org Xero Sync + Priority AI Queue + 25GB Storage", "$3,100 / mo labor saved", "Trading, logistics, e-commerce, multi-entity"],
        ["Accounting Agency / Multi-Client Tier", 2500.00, 699.00, "Up to 3,500 invoices/mo", "Multi-tenant portal (up to 10 company dashboards)", "$7,200 / mo labor saved", "Accounting firms managing multiple Xero books"],
    ]

    for row_idx, row_vals in enumerate(pricing_data, start=9):
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [2, 3]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            else:
                cell.font = font_data
                cell.alignment = align_center

    # Section 2: Multi-Company SaaS Revenue & Profit Projections (On 8GB Server)
    ws1.merge_cells("A13:G13")
    ws1["A13"] = "SAAS REVENUE & MARGIN MODEL (RUNNING ON SINGLE 8GB / 160-200GB SERVER)"
    ws1["A13"].font = font_sec_hdr
    ws1["A13"].fill = fill_sec_hdr
    
    headers_scale = ["Scale Stage", "Active Companies", "Avg Retainer / Co.", "Gross Monthly Revenue", "Total Server Cost (DO)", "Total AI Parsing Cost", "Net Monthly Profit"]
    for col_num, h in enumerate(headers_scale, 1):
        cell = ws1.cell(row=14, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    scale_data = [
        ["Current Launch Stage", 5, 199.00, "=B15*C15", 48.00, 0.75, "=D15-E15-F15"],
        ["Expanded SME Pool", 15, 199.00, "=B16*C16", 48.00, 2.25, "=D16-E16-F16"],
        ["Full 8GB Server Utilization", 35, 199.00, "=B17*C17", 58.00, 5.25, "=D17-E17-F17"],
        ["Future Scaled Cluster (16GB/1TB)", 100, 199.00, "=B18*C18", 120.00, 15.00, "=D18-E18-F18"],
    ]

    for row_idx, row_vals in enumerate(scale_data, start=15):
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [3, 4, 5, 6, 7]:
                cell.font = font_total if col_idx == 7 else font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
                if col_idx == 7:
                    cell.fill = fill_highlight
            else:
                cell.font = font_data
                cell.alignment = align_center

    # -------------------------------------------------------------
    # TAB 2: ITEMIZED INFRASTRUCTURE COSTS (8GB RAM / 100-200GB SSD)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Itemized 8GB Server Costs")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "ITEMIZED PRODUCTION INFRASTRUCTURE COSTS (8GB RAM / 160-200GB DISK)"
    ws2["A1"].font = font_title
    
    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Real-time production baseline supporting 20-50 active company mailboxes & dashboards"
    ws2["A2"].font = font_subtitle
    
    headers_infra = ["Category", "Component / Service Spec", "Capacity / Workload", "Billing Model", "DigitalOcean Route", "GCP Route"]
    for c_idx, h in enumerate(headers_infra, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    infra_items = [
        ["Server Compute (VPS)", "4 vCPU / 8 GB RAM / 160 GB NVMe SSD", "Handles 50+ IMAP sockets & Bull queue", "Fixed Monthly", 48.00, 68.50],
        ["Extra Storage Volume", "Optional +40GB NVMe Block Storage (Total 200GB)", "Holds 400,000+ invoice PDFs & DB", "Fixed ($0.10/GB)", 4.00, 8.00],
        ["Static Dedicated IPv4", "1 Dedicated Public IP for Webhooks/API", "Always-on public IP", "Monthly", 0.00, 3.65],
        ["Bandwidth / Egress", "High-speed network transfer for PDFs & UI", "5,000 GB (5TB) DO / Pay-per-GB GCP", "Included / Usage", 0.00, 8.00],
        ["AI LLM Parsing (Gemini)", "Google Gemini Flash-Lite (5,000 inv/mo)", "5 Million tokens/mo @ $0.075/1M", "Pay-as-you-go", 0.75, 0.75],
        ["Accounting Sync API", "Xero Developer API & OAuth2 (Multi-Tenant)", "60 RPM per tenant limit", "Free Platform Tier", 0.00, 0.00],
        ["Email Intake (IMAP)", "Real-time IMAP IDLE connections", "Direct to client's Gmail/O365 mailbox", "Zero Extra Cost", 0.00, 0.00],
        ["Domain, DNS & SSL", "Custom domain + Let's Encrypt / Cloudflare", "Automated HTTPS & DNS routing", "Amortized ($15/yr)", 1.25, 1.25],
        ["Automated Backups", "Daily SQLite DB snapshots to Offsite S3/R2", "Automated recovery snapshots", "Storage ($0.015/GB)", 2.00, 5.20],
        ["System Health & Alerts", "Slack Webhook alerts + Uptime Monitoring", "Instant crash/fail alerts", "Free Tier", 0.00, 0.00],
    ]

    for r_idx, row_vals in enumerate(infra_items, start=5):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx in [1, 2]:
                cell.font = font_data_bold if c_idx == 1 else font_data
                cell.alignment = align_left
            elif c_idx in [3, 4]:
                cell.font = font_data
                cell.alignment = align_center
            elif c_idx in [5, 6]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"

    # Total Row
    tot_row = len(infra_items) + 5
    ws2.cell(row=tot_row, column=1, value="TOTAL ESTIMATED DIRECT INFRASTRUCTURE COST (8GB / 200GB)").font = font_total
    ws2.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=4)
    for c in range(1, 5):
        ws2.cell(row=tot_row, column=c).fill = fill_total
        ws2.cell(row=tot_row, column=c).border = total_border

    c_do_tot = ws2.cell(row=tot_row, column=5, value=f"=SUM(E5:E{tot_row-1})")
    c_do_tot.font = font_total
    c_do_tot.fill = fill_total
    c_do_tot.border = total_border
    c_do_tot.alignment = align_right
    c_do_tot.number_format = "$#,##0.00"

    c_gcp_tot = ws2.cell(row=tot_row, column=6, value=f"=SUM(F5:F{tot_row-1})")
    c_gcp_tot.font = font_total
    c_gcp_tot.fill = fill_total
    c_gcp_tot.border = total_border
    c_gcp_tot.alignment = align_right
    c_gcp_tot.number_format = "$#,##0.00"

    # -------------------------------------------------------------
    # TAB 3: GCP VS DIGITALOCEAN COMPARISON (8GB / 200GB)
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="GCP vs DigitalOcean (8GB)")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "HOSTING PROVIDER COMPARISON: GCP VS DIGITALOCEAN (8GB RAM / 200GB DISK)"
    ws3["A1"].font = font_title
    
    headers_comp = ["Evaluation Dimension", "DigitalOcean (Droplet)", "Google Cloud Platform (GCP)", "Advantage / Winner", "Reasoning"]
    for c_idx, h in enumerate(headers_comp, 1):
        cell = ws3.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    comp_rows = [
        ["Monthly Base Spec (8GB / 4 vCPU)", "$48.00 / mo (Includes 160GB NVMe)", "$68.50 / mo (Compute Only)", "DigitalOcean", "DO provides 160GB local NVMe included in base price"],
        ["Storage Cost (to reach 200GB)", "+$4.00 / mo (40GB block volume)", "+$20.00 / mo (200GB Balanced SSD)", "DigitalOcean", "GCP charges separate rate for entire 200GB disk volume"],
        ["Static Public IPv4", "Free (Included with Droplet)", "$3.65 - $7.30 / mo", "DigitalOcean", "GCP charges continuous fee for reserved IPv4 address"],
        ["Bandwidth / Network Egress", "5,000 GB (5 TB) Included", "100 GB Free, then $0.08-$0.12/GB", "DigitalOcean", "Massive 5TB transfer prevents runaway bandwidth bills"],
        ["Multi-Tenant IMAP Socket Stability", "Excellent (Unfiltered TCP TLS)", "Requires firewall VPC tuning", "DigitalOcean", "Simpler persistent TCP socket handling for 50+ inboxes"],
        ["Price Predictability", "100% Fixed ($52.00 - $58.00 / mo)", "Variable ($95.00 - $115.00 / mo)", "DigitalOcean", "Fixed pricing protects your profit margins on client retainers"],
        ["Future Scaling Horizon", "1-click resize up to 192GB RAM / 10TB", "Global autoscaling & Kubernetes", "Tie / Context", "DO easily handles 1,000 companies; GCP for multi-region HA"],
        ["Final Recommendation", "HIGHLY RECOMMENDED (BEST VALUE)", "Alternative / Enterprise-Only", "DIGITALOCEAN", "Saves ~$600/year while delivering superior I/O disk speed"]
    ]

    for r_idx, row_vals in enumerate(comp_rows, start=5):
        fill_to_use = fill_highlight if "RECOMMENDED" in row_vals[1] or "DigitalOcean" in row_vals[3] else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx == 4:
                cell.font = font_data_bold
                cell.alignment = align_center
            else:
                cell.font = font_data
                cell.alignment = align_left

    # -------------------------------------------------------------
    # TAB 4: REAL-TIME ARCHITECTURE & SCALING CAPACITY
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Real-Time Architecture & Scale")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "REAL-TIME MULTI-COMPANY CAPACITY & RESOURCE UTILIZATION (8GB RAM)"
    ws4["A1"].font = font_title
    
    ws4.merge_cells("A2:E2")
    ws4["A2"] = "Technical sizing for simultaneous IMAP watchers, dashboard traffic & Bull queue workers"
    ws4["A2"].font = font_subtitle
    
    arch_headers = ["Subsystem / Component", "Per-Tenant Footprint", "Capacity on 8GB / 200GB Server", "Real-Life Behavior", "Bottleneck & Mitigation"]
    for c_idx, h in enumerate(arch_headers, 1):
        cell = ws4.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    arch_data = [
        ["Real-Time IMAP Watcher", "~25 KB RAM per open socket", "500+ active email sockets", "Keeps TLS IDLE open 24/7 with auto-reconnect", "Gmail drops idle sockets every 29m; auto-reconnect handled"],
        ["Bull / Redis Job Queue", "~50 MB base RAM for Redis", "10,000+ jobs in queue buffer", "Drains jobs sequentially with crash-recovery", "Disk-backed queue survives unexpected server reboots"],
        ["PDF Storage (Disk)", "~500 KB per invoice PDF", "400,000+ PDFs (200GB storage)", "Saved locally in data/users/{userId}/pdfs/", "Future scale: Sync to Cloudflare R2 / S3 when reaching 1M PDFs"],
        ["LLM Parser Concurrency", "15 RPM per tenant cap", "Up to 5 concurrent PDFs per email", "Extracts vendor, lines, tax, currency via Gemini", "Gemini key rotation prevents 429 quota exhaustion"],
        ["Live React Dashboard", "~1 MB per active user session", "100+ simultaneous web users", "Streams invoice reviews & signed PDF tokens", "Short-lived JWT tokens (5m TTL) protect invoice PDFs"],
        ["SQLite Database Engine", "~100 MB RAM for WAL cache", "Millions of invoice rows in app.db", "Fast ACID transactions with WAL mode", "Automated daily sqlite3 .backup snapshots to offsite storage"]
    ]

    for r_idx, row_vals in enumerate(arch_data, start=5):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.font = font_data_bold if c_idx == 1 else font_data
            if c_idx == 1:
                cell.alignment = align_left
            elif c_idx == 3:
                cell.alignment = align_center
                cell.fill = fill_highlight
            else:
                cell.alignment = align_left

    # Column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str += "    "
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
    wb.save(XLSX_PATH)
    print(f"Updated Excel report created at: {XLSX_PATH}")


# ==============================================================================
# 2. PDF REPORT GENERATOR (Multi-Company & 8GB / 200GB Accounting Format)
# ==============================================================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Header (pages 2+)
        if self._pageNumber > 1:
            self.drawString(54, 11 * inch - 36, "XERO INVOICE AUTOMATION — MULTI-TENANT PRODUCTION COST REPORT")
            self.drawRightString(8.5 * inch - 54, 11 * inch - 36, "CONFIDENTIAL")
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(54, 11 * inch - 42, 8.5 * inch - 54, 11 * inch - 42)
            
        # Footer (all pages)
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(54, 46, 8.5 * inch - 54, 46)
        
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawString(54, 32, "Prepared for Multi-Company SaaS & Infrastructure Planning | August 2026")
        self.drawRightString(8.5 * inch - 54, 32, page_str)
        self.restoreState()


def create_pdf_report():
    doc = SimpleDocTemplate(
        PDF_PATH,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    c_navy = colors.HexColor("#1E3A8A")
    c_slate = colors.HexColor("#334155")
    c_body = colors.HexColor("#1E293B")
    c_teal = colors.HexColor("#0D9488")
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=18, leading=22,
        textColor=c_navy, spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'],
        fontName='Helvetica-Oblique', fontSize=9.5, leading=13,
        textColor=colors.HexColor("#64748B"), spaceAfter=12
    )
    h1_style = ParagraphStyle(
        'SecH1', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=11, leading=15,
        textColor=c_navy, spaceBefore=10, spaceAfter=5, keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyDark', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8.5, leading=12,
        textColor=c_body, spaceAfter=5
    )
    th_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, leading=10,
        textColor=colors.white, alignment=1
    )
    td_left = ParagraphStyle(
        'TableCellLeft', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.5, leading=9.5,
        textColor=c_body
    )
    td_bold = ParagraphStyle(
        'TableCellBold', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
        textColor=c_body
    )
    td_right = ParagraphStyle(
        'TableCellRight', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.5, leading=9.5,
        textColor=c_body, alignment=2
    )
    td_center = ParagraphStyle(
        'TableCellCenter', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.5, leading=9.5,
        textColor=c_body, alignment=1
    )
    
    elements = []
    
    # ── HEADER BLOCK ──
    elements.append(Paragraph("XERO MULTI-TENANT INVOICE AUTOMATION SYSTEM", title_style))
    elements.append(Paragraph("Cost Estimation, Multi-Company Capacity Analysis (8GB RAM / 100-200GB Storage) & Pricing Model", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=c_navy, spaceAfter=10))
    
    # ── EXECUTIVE SUMMARY / METADATA ──
    meta_table_data = [
        [Paragraph("<b>Target Server Spec:</b> 8 GB RAM | 160–200 GB NVMe SSD | 4 vCPU", td_left), Paragraph("<b>Target Architecture:</b> Multi-Tenant Real-Time IMAP + Xero OAuth", td_left)],
        [Paragraph("<b>Estimated Active Capacity:</b> 20 to 50 Connected Companies", td_left), Paragraph("<b>Target Monthly Infra Cost (DO):</b> ~$52.00 – $58.00 / month total", td_left)],
        [Paragraph("<b>Storage Capacity:</b> 400,000+ Invoice PDFs + WAL Database", td_left), Paragraph("<b>Net Profit Margin:</b> > 95% at 5+ onboarded companies", td_left)]
    ]
    t_meta = Table(meta_table_data, colWidths=[270, 234])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 1: HOW THE MULTI-COMPANY REAL-TIME SYSTEM WORKS ──
    elements.append(Paragraph("1. SYSTEM ARCHITECTURE & REAL-TIME MULTI-COMPANY MECHANICS", h1_style))
    elements.append(Paragraph(
        "Each onboarded company operates as an isolated tenant inside the system with dedicated credentials, storage, and live synchronization:",
        body_style
    ))
    
    mech_tbl = [
        [
            Paragraph("System Layer", th_style),
            Paragraph("How It Operates (Real-Life Real-Time Pipeline)", th_style),
            Paragraph("8GB RAM / 200GB Sizing Impact", th_style),
        ],
        [
            Paragraph("<b>Real-Time Email Watcher (IMAP IDLE)</b>", td_left),
            Paragraph("Holds a persistent 24/7 TLS connection to each company's mailbox (Gmail/O365). Incoming invoices trigger instant instant notifications with 60s fallback poll.", td_left),
            Paragraph("<b>~25 KB RAM per company.</b> 8GB RAM effortlessly runs 100+ concurrent mailbox connections.", td_left),
        ],
        [
            Paragraph("<b>Queue & PDF Extraction Worker</b>", td_left),
            Paragraph("Extracts PDFs and pushes them to a crash-safe queue. Parses up to 5 PDFs per email concurrently, extracting vendor, dates, line items, taxes, and amounts.", td_left),
            Paragraph("<b>Disk-backed queue</b> survives server reboots; 8GB RAM prevents out-of-memory spikes.", td_left),
        ],
        [
            Paragraph("<b>AI Invoice Parsing (Gemini)</b>", td_left),
            Paragraph("Structured JSON extraction via Google Gemini Flash-Lite with 15 RPM per-tenant rate limit and automatic API key rotation on quota exhaustion.", td_left),
            Paragraph("<b>~$0.00015 per invoice</b> (<$1.00/mo across thousands of invoices).", td_left),
        ],
        [
            Paragraph("<b>Live Dashboard & Review Portal</b>", td_left),
            Paragraph("Company staff can log in to view real-time sync status, review draft bills, edit fields, preview PDFs via signed 5m tokens, and trigger manual rescans.", td_left),
            Paragraph("<b>100+ simultaneous web users</b> supported on fast React SPA + Node backend.", td_left),
        ],
        [
            Paragraph("<b>PDF Storage & Database Archival</b>", td_left),
            Paragraph("Stores original vendor PDFs and transaction logs in tenant-isolated directories (data/users/{userId}/pdfs/) + SQLite WAL database.", td_left),
            Paragraph("<b>200 GB SSD holds ~400,000 PDFs</b>, covering years of multi-company operations.", td_left),
        ]
    ]
    t_mech = Table(mech_tbl, colWidths=[120, 234, 150])
    t_mech.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_mech)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 2: 8GB / 200GB INFRASTRUCTURE ITEMIZED COSTS ──
    elements.append(Paragraph("2. ITEMIZED PRODUCTION INFRASTRUCTURE COSTS (8GB RAM / 160-200GB SSD)", h1_style))
    
    infra_tbl_data = [
        [
            Paragraph("Component / Service", th_style),
            Paragraph("Specification & Role", th_style),
            Paragraph("DigitalOcean Route", th_style),
            Paragraph("GCP (Google Cloud)", th_style),
        ],
        [
            Paragraph("<b>Cloud Server (VPS)</b>", td_left),
            Paragraph("4 vCPU, 8 GB RAM, 160 GB NVMe SSD (5TB Transfer)", td_left),
            Paragraph("<b>$48.00 / mo</b>", td_right),
            Paragraph("$68.50 / mo", td_right),
        ],
        [
            Paragraph("<b>Extra NVMe Storage Volume</b>", td_left),
            Paragraph("+40 GB NVMe Block Storage (Reaches 200GB Total)", td_left),
            Paragraph("<b>$4.00 / mo</b>", td_right),
            Paragraph("$8.00 / mo", td_right),
        ],
        [
            Paragraph("<b>Dedicated Static IPv4</b>", td_left),
            Paragraph("1 Public IP for Xero Webhooks & HTTPS Server", td_left),
            Paragraph("<b>$0.00</b> (Included)", td_right),
            Paragraph("$3.65 / mo", td_right),
        ],
        [
            Paragraph("<b>Network Transfer / Egress</b>", td_left),
            Paragraph("5,000 GB (5 TB) transfer included on DO", td_left),
            Paragraph("<b>$0.00</b> (Included)", td_right),
            Paragraph("~$8.00 / mo", td_right),
        ],
        [
            Paragraph("<b>Gemini AI Token Parsing</b>", td_left),
            Paragraph("5,000 invoices/mo (~5M tokens @ $0.075/1M)", td_left),
            Paragraph("<b>$0.75 / mo</b>", td_right),
            Paragraph("$0.75 / mo", td_right),
        ],
        [
            Paragraph("<b>Xero Platform API</b>", td_left),
            Paragraph("Multi-tenant Custom Connection & OAuth2", td_left),
            Paragraph("<b>$0.00</b> (Free Tier)", td_right),
            Paragraph("$0.00 (Free Tier)", td_right),
        ],
        [
            Paragraph("<b>Domain & SSL Security</b>", td_left),
            Paragraph("Custom domain (.com) + Let's Encrypt / Cloudflare", td_left),
            Paragraph("<b>$1.25 / mo</b>", td_right),
            Paragraph("$1.25 / mo", td_right),
        ],
        [
            Paragraph("<b>Offsite Snapshot Backups</b>", td_left),
            Paragraph("Automated daily SQLite backup to Cloudflare R2 / S3", td_left),
            Paragraph("<b>$2.00 / mo</b>", td_right),
            Paragraph("$5.20 / mo", td_right),
        ],
        [
            Paragraph("<b>TOTAL REAL MONTHLY COST</b>", td_bold),
            Paragraph("<b>Complete 24/7 multi-company hosting cluster</b>", td_bold),
            Paragraph("<b>$56.00 / month</b>", td_bold),
            Paragraph("<b>$95.35 / month</b>", td_bold),
        ]
    ]
    t_infra = Table(infra_tbl_data, colWidths=[130, 204, 85, 85])
    t_infra.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#DBEAFE")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_infra)
    elements.append(Spacer(1, 10))
    
    # ── PAGE BREAK ──
    elements.append(PageBreak())
    
    # ── SECTION 3: MULTI-COMPANY SAAS PROFITABILITY & PRICING ──
    elements.append(Paragraph("3. HOW TO CHARGE EACH COMPANY (SAAS REVENUE & PROFIT MODEL)", h1_style))
    elements.append(Paragraph(
        "Because one 8GB / 200GB server ($56/mo) easily handles 20 to 50 companies, your server expense becomes a tiny fraction of your monthly revenue:",
        body_style
    ))
    
    scale_tbl_data = [
        [
            Paragraph("Onboarded Scale", th_style),
            Paragraph("Monthly Fee / Co.", th_style),
            Paragraph("Gross Monthly Revenue", th_style),
            Paragraph("Total Server Cost", th_style),
            Paragraph("Net Monthly Profit", th_style),
            Paragraph("Gross Profit Margin", th_style),
        ],
        [
            Paragraph("<b>3 Companies (Launch)</b>", td_left),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$597.00 / mo", td_right),
            Paragraph("$56.00 / mo", td_right),
            Paragraph("<b>$541.00 / mo</b>", td_right),
            Paragraph("<b>90.6%</b>", td_center),
        ],
        [
            Paragraph("<b>10 Companies (Established)</b>", td_left),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$1,990.00 / mo", td_right),
            Paragraph("$56.00 / mo", td_right),
            Paragraph("<b>$1,934.00 / mo</b>", td_right),
            Paragraph("<b>97.2%</b>", td_center),
        ],
        [
            Paragraph("<b>25 Companies (Scaling)</b>", td_left),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$4,975.00 / mo", td_right),
            Paragraph("$56.00 / mo", td_right),
            Paragraph("<b>$4,919.00 / mo</b>", td_right),
            Paragraph("<b>98.8%</b>", td_center),
        ],
        [
            Paragraph("<b>50 Companies (Capacity)</b>", td_left),
            Paragraph("$199.00 / mo", td_right),
            Paragraph("$9,950.00 / mo", td_right),
            Paragraph("$68.00 / mo", td_right),
            Paragraph("<b>$9,882.00 / mo</b>", td_right),
            Paragraph("<b>99.3%</b>", td_center),
        ]
    ]
    t_scale = Table(scale_tbl_data, colWidths=[120, 75, 85, 74, 85, 65])
    t_scale.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (4,1), (4,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t_scale)
    elements.append(Spacer(1, 10))
    
    # ── SECTION 4: GCP VS DIGITALOCEAN (8GB COMPARISON) ──
    elements.append(Paragraph("4. GCP VS. DIGITALOCEAN COMPARISON (8GB RAM / 200GB DISK)", h1_style))
    
    gcp_do_tbl = [
        [
            Paragraph("Dimension", th_style),
            Paragraph("DigitalOcean (Droplet)", th_style),
            Paragraph("Google Cloud Platform (GCP)", th_style),
            Paragraph("Strategic Recommendation", th_style),
        ],
        [
            Paragraph("<b>Base Monthly Price</b>", td_left),
            Paragraph("$48.00 / mo (4 vCPU, 8GB, 160GB SSD)", td_left),
            Paragraph("$68.50 / mo (2 vCPU, 8GB compute only)", td_left),
            Paragraph("<b>DO is 40% cheaper + gives double CPU</b>", td_left),
        ],
        [
            Paragraph("<b>Disk Storage (200GB)</b>", td_left),
            Paragraph("160GB NVMe free + $4/mo for +40GB", td_left),
            Paragraph("+$20.00/mo (200GB Balanced SSD)", td_left),
            Paragraph("<b>DO saves $16/mo on high-speed SSD</b>", td_left),
        ],
        [
            Paragraph("<b>Static IPv4 & Bandwidth</b>", td_left),
            Paragraph("Free static IP + 5,000 GB (5TB) transfer", td_left),
            Paragraph("+$3.65/mo IP + $0.08-$0.12/GB egress", td_left),
            Paragraph("<b>DO eliminates unexpected egress bills</b>", td_left),
        ],
        [
            Paragraph("<b>IMAP Connection Stability</b>", td_left),
            Paragraph("Clean direct TCP TLS socket handling", td_left),
            Paragraph("Requires VPC egress / firewall setup", td_left),
            Paragraph("<b>DO requires zero DevOps maintenance</b>", td_left),
        ],
        [
            Paragraph("<b>Total Monthly Cost</b>", td_bold),
            Paragraph("<b>$52.00 – $58.00 / month flat</b>", td_bold),
            Paragraph("<b>$95.00 – $115.00 / month</b>", td_bold),
            Paragraph("<b>DigitalOcean is the clear winner</b>", td_bold),
        ]
    ]
    t_gcp = Table(gcp_do_tbl, colWidths=[110, 130, 130, 134])
    t_gcp.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#DBEAFE")),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t_gcp)
    elements.append(Spacer(1, 10))
    
    # ── SECTION 5: SUMMARY & NEXT STEPS ──
    elements.append(Paragraph("5. SUMMARY FOR CLIENT PRESENTATION", h1_style))
    elements.append(Paragraph(
        "• <b>Real Hosting Cost:</b> The complete production server with <b>8 GB RAM and 200 GB NVMe Storage</b> costs <b>~$56.00 / month on DigitalOcean</b>.<br/>"
        "• <b>Multi-Company Capacity:</b> This single server easily supports <b>20 to 50 active companies</b> with live IMAP email watchers and real-time review dashboards.<br/>"
        "• <b>Client Pricing:</b> Quote each company <b>$199.00 / month</b> (or $150–$250/mo depending on invoice volume). The client saves over <b>$1,000/month in manual bookkeeper hours</b>, delivering an immediate <b>300%+ ROI</b> to their business.",
        body_style
    ))
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    print(f"Updated PDF report created at: {PDF_PATH}")


# ==============================================================================
# 3. MARKDOWN COMPANION GENERATOR (Multi-Company & 8GB / 200GB Scale)
# ==============================================================================
def create_md_report():
    md_content = """# Xero Multi-Tenant Invoice Automation — Cost Estimation & System Capacity Report

**Document Type:** Multi-Company Commercial Feasibility & Production Capacity Report  
**Target Specification:** 8 GB RAM | 160–200 GB NVMe SSD | 4 vCPU  
**Architecture:** Multi-Tenant Node.js + SQLite (WAL Mode) + Bull/Redis Queue + Google Gemini AI  
**Generated Files:**
- **Excel Spreadsheet (Formulas & Tiers):** [`XERO_AUTOMATION_COST_ESTIMATION_REPORT.xlsx`](file:///Users/weikangten/Desktop/xero-invoice-app-master/report/XERO_AUTOMATION_COST_ESTIMATION_REPORT.xlsx)
- **Formal PDF Report (Accounting Grade):** [`XERO_AUTOMATION_COST_ESTIMATION_REPORT.pdf`](file:///Users/weikangten/Desktop/xero-invoice-app-master/report/XERO_AUTOMATION_COST_ESTIMATION_REPORT.pdf)

---

## 1. System Architecture & Real-Time Production Mechanics

In real production, each company connects independently to the platform:

```
┌───────────────────────────────────────────────────────────────────────────────────┐
│                      REAL-TIME MULTI-TENANT ARCHITECTURE                          │
├───────────────────────────────────────────────────────────────────────────────────┤
│ 1. REAL-TIME IMAP WATCHER (Per Company):                                          │
│    • Holds a persistent 24/7 TLS connection to company inboxes (Gmail / O365).    │
│    • Uses IMAP IDLE for instant zero-latency email arrival + 60s fallback poll.  │
│    • Memory footprint: ~25 KB per connected company (8GB RAM holds 500+ sockets). │
│                                                                                   │
│ 2. QUEUE & PARSING WORKER (Bull + Disk-Backed Queue):                             │
│    • Extracts PDF attachments (up to 5 concurrent PDFs per email).                │
│    • Gemini Flash-Lite AI parses vendor, date, line items, taxes, currency.       │
│    • Rate limited (15 RPM per tenant) with automatic API key rotation on quota.   │
│                                                                                   │
│ 3. MULTI-TENANT DASHBOARD & REVIEW PORTAL:                                        │
│    • Company staff log in to review draft bills, edit fields, and approve.        │
│    • Secure signed PDF tokens (5 min TTL) stream invoices directly in browser.    │
│                                                                                   │
│ 4. XERO OAUTH2 / CUSTOM CONNECTION:                                               │
│    • Sequential submission queue (1.5s gap) strictly stays inside Xero 60 RPM.    │
│    • Automatic token refresh and base-currency detection.                         │
│                                                                                   │
│ 5. STORAGE & BACKUPS (200 GB SSD):                                                │
│    • Stores 400,000+ invoice PDFs in data/users/{userId}/pdfs/.                   │
│    • High-performance SQLite database with Write-Ahead Logging (WAL mode).        │
│    • Automated daily snapshot backups to offsite S3 / Cloudflare R2 storage.      │
└───────────────────────────────────────────────────────────────────────────────────┘
```

---

## 2. Real Production Infrastructure Costs (8GB RAM / 160–200GB Storage)

| Infrastructure Component | Specification / Capacity | DigitalOcean Route | GCP (Compute Engine) |
|---|---|---|---|
| **Base Cloud VPS** | 4 vCPU / 8 GB RAM / 160 GB NVMe SSD | **$48.00 / mo** | $68.50 / mo (2 vCPU) |
| **Storage Expansion (to 200GB)** | +40 GB NVMe Block Storage Volume | **$4.00 / mo** | $20.00 / mo (200GB SSD) |
| **Dedicated Static IPv4** | 1 Dedicated Public IP for Webhooks & SSL | **$0.00** (Included) | $3.65 / mo |
| **Bandwidth / Egress** | 5,000 GB (5 TB) Included on DigitalOcean | **$0.00** (Included) | ~$8.00 / mo |
| **AI LLM Invoice Parsing** | Google Gemini Flash-Lite (~5,000 invoices/mo) | **$0.75 / mo** | $0.75 / mo |
| **Xero Platform API** | Multi-tenant OAuth2 & Custom Connections | **$0.00** (Free Developer) | $0.00 (Free Developer) |
| **Domain & SSL Certificate** | Custom Domain (.com) + Let's Encrypt SSL | **$1.25 / mo** | $1.25 / mo |
| **Offsite Snapshot Backups** | Daily SQLite DB snapshots (Cloudflare R2 / S3) | **$2.00 / mo** | $5.20 / mo |
| **TOTAL REAL SERVER EXPENSE** | **Full 24/7 cluster for 20–50 companies** | **~$56.00 / month** | **~$107.35 / month** |

---

## 3. DigitalOcean vs. Google Cloud Platform (GCP) Comparison

| Evaluation Dimension | DigitalOcean Droplet | Google Cloud Platform (GCP) | Final Verdict |
|---|---|---|---|
| **Monthly Pricing (8GB / 200GB)** | **~$52.00 – $58.00 / mo flat** | **~$95.00 – $115.00 / mo variable** | **DigitalOcean saves ~$600/year** |
| **CPU Performance** | **4 Dedicated/Shared vCPUs** | 2 vCPUs on baseline `e2-standard-2` | **DO gives 2x CPU cores for PDF parsing** |
| **Disk Storage (NVMe SSD)** | 160GB NVMe included + cheap block volume | Charged per GB/mo continuously | **DO provides faster local SQLite I/O** |
| **Bandwidth & Static IP** | 5,000 GB (5 TB) + Free Static IP | 100 GB free, then pay per GB + IP fee | **DO eliminates billing surprise risk** |
| **DevOps Maintenance** | Minimal (Clean UI, simple firewall & SSH) | High (IAM policies, VPC routing, alarms) | **DO saves 10+ hours in dev setup** |

> **Recommendation:** **DigitalOcean is the optimal choice** for this single-instance Node.js + SQLite multi-tenant server.

---

## 4. Multi-Company SaaS Economics: How to Price & Scale

Because one **$56/mo server** handles **20 to 50 companies**, your profit margins are over **95%**:

| Stage | Connected Companies | Retainer per Company | Monthly Revenue | Total Server Cost | Net Monthly Profit | Gross Margin |
|---|---|---|---|---|---|---|
| **Launch Stage** | **3 Companies** | $199.00 / mo | **$597.00 / mo** | $56.00 / mo | **+$541.00 / mo** | **90.6%** |
| **Growth Stage** | **10 Companies** | $199.00 / mo | **$1,990.00 / mo** | $56.00 / mo | **+$1,934.00 / mo** | **97.2%** |
| **Established Pool** | **25 Companies** | $199.00 / mo | **$4,975.00 / mo** | $56.00 / mo | **+$4,919.00 / mo** | **98.8%** |
| **Server Full Capacity** | **50 Companies** | $199.00 / mo | **$9,950.00 / mo** | $68.00 / mo | **+$9,882.00 / mo** | **99.3%** |

---

## 5. Client ROI & Value Proposition

When selling this service to a company:
* **Manual Data Entry Cost:** Entering 500 invoices/month by hand takes **41.6 hours of bookkeeper time** = **~$1,040 / month**.
* **Automated System Cost:** Fast 30-second review on the dashboard = **$104 labor + $199 retainer** = **$303 / month**.
* **Client Direct Savings:** The client saves **$738.50 every month ($8,860/year)** while gaining real-time Xero synchronization and eliminating human data entry errors.
"""
    with open(MD_PATH, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Updated Markdown report created at: {MD_PATH}")


if __name__ == "__main__":
    create_excel_report()
    create_pdf_report()
    create_md_report()

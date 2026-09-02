import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

REPORT_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report"
ARCH_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report/arch_recommendation"

# ==============================================================================
# 1. EXCEL REPORT GENERATOR (Realistic & Competitive Pricing)
# ==============================================================================
def create_excel_report(out_path):
    wb = openpyxl.Workbook()
    
    NAVY_DARK = "1E3A8A"
    NAVY_LIGHT = "DBEAFE"
    SLATE_DARK = "334155"
    SLATE_LIGHT = "F8FAFC"
    GREEN_ACCENT = "0D9488"
    GREEN_LIGHT = "CCFBF1"
    GRAY_TEXT = "64748B"
    WHITE = "FFFFFF"
    
    font_title = Font(name="Calibri", size=16, bold=True, color=NAVY_DARK)
    font_subtitle = Font(name="Calibri", size=11, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=12, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=11, color="1E293B")
    font_data_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_total = Font(name="Calibri", size=11, bold=True, color=NAVY_DARK)
    
    fill_sec_hdr = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=SLATE_DARK, end_color=SLATE_DARK, fill_type="solid")
    fill_alt_row = PatternFill(start_color=SLATE_LIGHT, end_color=SLATE_LIGHT, fill_type="solid")
    fill_highlight = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    fill_total = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    total_border = Border(
        top=Side(style='thin', color='1E3A8A'),
        bottom=Side(style='double', color='1E3A8A'),
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    # -------------------------------------------------------------
    # TAB 1: REALISTIC CLIENT PRICING PACKAGES
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Realistic Client Pricing"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "XERO INVOICE AUTOMATION — REALISTIC COST & PRICING PROPOSAL"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_left
    
    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Competitive, High-Conversion Pricing Model with Zero-Hidden-Cost Guarantee"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = align_left
    
    # Metadata
    ws1["A4"] = "Target Hardware:"
    ws1["B4"] = "8 GB RAM | 160-200 GB NVMe SSD"
    ws1["A5"] = "Gemini AI Cost:"
    ws1["B5"] = "$0.00 (Google AI Free Tier: 15 RPM / 1,500 req/day)"
    ws1["D4"] = "Setup Time Required:"
    ws1["E4"] = "15 to 30 Minutes (Self-serve / Guided)"
    ws1["D5"] = "Base Currency:"
    ws1["E5"] = "USD ($)"
    for r in [4, 5]:
        ws1[f"A{r}"].font = font_data_bold
        ws1[f"B{r}"].font = font_data
        ws1[f"D{r}"].font = font_data_bold
        ws1[f"E{r}"].font = font_data

    # Section 1: Realistic Commercial Packages
    ws1.merge_cells("A7:G7")
    ws1["A7"] = "PRACTICAL & COMPETITIVE CLIENT PACKAGES (SME FRIENDLY)"
    ws1["A7"].font = font_sec_hdr
    ws1["A7"].fill = fill_sec_hdr

    headers_pricing = [
        "Package Tier", "One-Time Setup Fee", "Monthly Retainer", 
        "Included Volume", "Real Server Cost", "Net Monthly Margin", "Target Client Profile"
    ]
    for col_num, h in enumerate(headers_pricing, 1):
        cell = ws1.cell(row=8, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    pricing_data = [
        ["Starter SME Plan (Budget Friendly)", 150.00, 99.00, "Up to 150 invoices/mo", 3.00, 96.00, "Small business with 20–150 monthly bills"],
        ["Standard Business Plan (Recommended)", 250.00, 149.00, "Up to 500 invoices/mo", 5.00, 144.00, "Standard SME with regular daily supplier bills"],
        ["Pro / High-Volume Plan", 350.00, 249.00, "Up to 1,500 invoices/mo", 10.00, 239.00, "Trading, logistics, e-commerce with heavy volume"],
    ]

    for row_idx, row_vals in enumerate(pricing_data, start=9):
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [2, 3, 5, 6]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
                if col_idx == 6:
                    cell.fill = fill_highlight
                    cell.font = font_total
            else:
                cell.font = font_data
                cell.alignment = align_center if col_idx == 4 else align_left

    # Section 2: Real Setup Effort Breakdown (15-30 mins)
    ws1.merge_cells("A13:G13")
    ws1["A13"] = "REALISTIC SETUP EFFORT & TRANSPARENT TIME AUDIT"
    ws1["A13"].font = font_sec_hdr
    ws1["A13"].fill = fill_sec_hdr

    headers_setup = ["Setup Step", "Actual Time Needed", "Vendor Fee", "Why It Is Fast", "Client Setup Charge"]
    for col_num, h in enumerate(headers_setup, 1):
        cell = ws1.cell(row=14, column=col_num, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    setup_data = [
        ["1. Connect Xero OAuth2 Account", "5 - 10 mins", 0.00, "1-click OAuth button in Setup UI (auto-fetches Tenant ID & org currency)", 100.00],
        ["2. Connect IMAP Mailbox (Gmail / M365)", "5 - 10 mins", 0.00, "Generate 16-character App Password, enter into Setup page, click Test Connection", 75.00],
        ["3. Register Free Google Gemini API Key", "3 - 5 mins", 0.00, "Generate free key at aistudio.google.com (15 RPM / 1,500 req/day). Zero credit card needed", 50.00],
        ["4. Send 1 Test Email & Verification", "2 - 5 mins", 0.00, "Send sample supplier invoice, check live dashboard for draft creation in Xero", 25.00],
    ]

    for row_idx, row_vals in enumerate(setup_data, start=15):
        fill_to_use = fill_alt_row if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if col_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif col_idx in [3, 5]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            elif col_idx == 2:
                cell.font = font_data
                cell.alignment = align_center
            else:
                cell.font = font_data
                cell.alignment = align_left

    tot_setup_row = len(setup_data) + 15
    ws1.cell(row=tot_setup_row, column=1, value="TOTAL SETUP INVESTMENT").font = font_total
    ws1.cell(row=tot_setup_row, column=2, value="15 - 30 Minutes").font = font_total
    ws1.cell(row=tot_setup_row, column=2).alignment = align_center
    ws1.cell(row=tot_setup_row, column=3, value=0.00).number_format = "$#,##0.00"
    ws1.cell(row=tot_setup_row, column=4, value="Pre-built automated pipeline in codebase").font = font_data
    ws1.cell(row=tot_setup_row, column=5, value=f"=SUM(E15:E{tot_setup_row-1})").number_format = "$#,##0.00"
    
    for c in range(1, 6):
        ws1.cell(row=tot_setup_row, column=c).fill = fill_total
        ws1.cell(row=tot_setup_row, column=c).border = total_border
        if c in [3, 5]:
            ws1.cell(row=tot_setup_row, column=c).font = font_total
            ws1.cell(row=tot_setup_row, column=c).alignment = align_right

    # -------------------------------------------------------------
    # TAB 2: ITEMIZED CLOUD & API OPERATING COSTS
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Itemized Costs & API")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F1")
    ws2["A1"] = "ITEMIZED HARD INFRASTRUCTURE & VENDOR OPERATING COSTS"
    ws2["A1"].font = font_title
    
    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Real server costs on DigitalOcean vs GCP supporting 20-50 connected companies"
    ws2["A2"].font = font_subtitle
    
    headers_infra = ["Category", "Line Item / Provider", "Base Specs / Quota", "Pricing Basis", "DigitalOcean Route", "GCP Route"]
    for c_idx, h in enumerate(headers_infra, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    infra_items = [
        ["Cloud Server (VPS)", "4 vCPU / 8 GB RAM / 160 GB NVMe SSD", "Handles 20-50 active IMAP sockets", "Fixed Monthly", 48.00, 68.50],
        ["NVMe Storage Volume", "+40 GB block storage (200GB Total SSD)", "Holds 400,000+ invoice PDFs & DB", "Fixed ($0.10/GB)", 4.00, 8.00],
        ["Dedicated Static IPv4", "1 Public IP for Xero Webhooks/SSH", "Fixed IP address", "Monthly", 0.00, 3.65],
        ["Bandwidth / Transfer", "High-speed data egress for PDFs & UI", "5,000 GB (5TB) DO / Pay-per-GB GCP", "Included / Usage", 0.00, 8.00],
        ["AI LLM Parsing (Gemini)", "Google Gemini Free Tier (aistudio.google.com)", "15 RPM / 1,500 req/day quota", "100% Free Forever", 0.00, 0.00],
        ["Gemini API Registration", "Google AI Studio Account Creation", "No payment method needed", "Free Registration", 0.00, 0.00],
        ["Xero Platform API", "OAuth2 / Custom Connection (60 RPM/org)", "Multi-tenant developer portal", "Free Developer Tier", 0.00, 0.00],
        ["Email Intake (IMAP)", "Client IMAP / Gmail App Password", "Direct connection to existing inbox", "Existing Account", 0.00, 0.00],
        ["Domain, DNS & SSL", "Custom Domain Name (.com / .app)", "Automated Let's Encrypt SSL", "Amortized ($15/yr)", 1.25, 1.25],
        ["Automated Backups", "Daily SQLite DB snapshots (Cloudflare R2/S3)", "Automated disaster recovery", "Storage ($0.015/GB)", 2.00, 5.20],
        ["Monitoring & Alerts", "Slack Webhook alerts + Uptime Check", "Instant crash/fail alerts", "Free Tier", 0.00, 0.00]
    ]

    for r_idx, row_vals in enumerate(infra_items, start=5):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx in [1, 2]:
                cell.font = font_data_bold if c_idx == 1 else font_data
                cell.alignment = align_left
            elif c_idx in [3, 4]:
                cell.font = font_data
                cell.alignment = align_center
            elif c_idx in [5, 6]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"

    tot_infra_row = len(infra_items) + 5
    ws2.cell(row=tot_infra_row, column=1, value="TOTAL ESTIMATED DIRECT MONTHLY INFRASTRUCTURE COST").font = font_total
    ws2.merge_cells(start_row=tot_infra_row, start_column=1, end_row=tot_infra_row, end_column=4)
    for c in range(1, 5):
        ws2.cell(row=tot_infra_row, column=c).fill = fill_total
        ws2.cell(row=tot_infra_row, column=c).border = total_border

    c_do_tot = ws2.cell(row=tot_infra_row, column=5, value=f"=SUM(E5:E{tot_infra_row-1})")
    c_do_tot.font = font_total
    c_do_tot.fill = fill_total
    c_do_tot.border = total_border
    c_do_tot.alignment = align_right
    c_do_tot.number_format = "$#,##0.00"

    c_gcp_tot = ws2.cell(row=tot_infra_row, column=6, value=f"=SUM(F5:F{tot_infra_row-1})")
    c_gcp_tot.font = font_total
    c_gcp_tot.fill = fill_total
    c_gcp_tot.border = total_border
    c_gcp_tot.alignment = align_right
    c_gcp_tot.number_format = "$#,##0.00"

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str += "    "
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
    wb.save(out_path)
    print(f"Realistic Excel report created at: {out_path}")


# ==============================================================================
# 2. PDF REPORT GENERATOR (Realistic Pricing & Zero Hidden Costs)
# ==============================================================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Header (pages 2+)
        if self._pageNumber > 1:
            self.drawString(54, 11 * inch - 36, "XERO INVOICE AUTOMATION — COST ESTIMATION REPORT")
            self.drawRightString(8.5 * inch - 54, 11 * inch - 36, "CONFIDENTIAL")
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(54, 11 * inch - 42, 8.5 * inch - 54, 11 * inch - 42)
            
        # Footer
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(54, 46, 8.5 * inch - 54, 46)
        
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawString(54, 32, "Commercial Cost Estimation & Pricing Report | August 2026")
        self.drawRightString(8.5 * inch - 54, 32, page_str)
        self.restoreState()


def create_pdf_report(out_path):
    doc = SimpleDocTemplate(
        out_path,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    c_navy = colors.HexColor("#1E3A8A")
    c_slate = colors.HexColor("#334155")
    c_body = colors.HexColor("#1E293B")
    c_teal = colors.HexColor("#0D9488")
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=17, leading=21,
        textColor=c_navy, spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'],
        fontName='Helvetica-Oblique', fontSize=9, leading=12,
        textColor=colors.HexColor("#64748B"), spaceAfter=10
    )
    h1_style = ParagraphStyle(
        'SecH1', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10.5, leading=14,
        textColor=c_navy, spaceBefore=9, spaceAfter=4, keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyDark', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, leading=11,
        textColor=c_body, spaceAfter=4
    )
    th_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
        textColor=colors.white, alignment=1
    )
    td_left = ParagraphStyle(
        'TableCellLeft', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_bold = ParagraphStyle(
        'TableCellBold', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_right = ParagraphStyle(
        'TableCellRight', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=2
    )
    td_center = ParagraphStyle(
        'TableCellCenter', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=1
    )
    
    elements = []
    
    # ── HEADER ──
    elements.append(Paragraph("XERO INVOICE AUTOMATION — COST & PRICING REPORT", title_style))
    elements.append(Paragraph("Realistic SME Pricing, Transparent Setup Effort (15–30 Mins) & Zero Hidden Gemini API Cost", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=c_navy, spaceAfter=8))
    
    # ── METADATA ──
    meta_table_data = [
        [Paragraph("<b>Target Stack:</b> Multi-Tenant Node.js + SQLite WAL + React + Bull", td_left), Paragraph("<b>Date:</b> August 2026", td_left)],
        [Paragraph("<b>Gemini AI API Cost:</b> $0.00 (Google AI Studio Free Tier: 15 RPM)", td_left), Paragraph("<b>Server Cost:</b> ~$55.25 / mo on DigitalOcean (8GB/200GB)", td_left)],
        [Paragraph("<b>Setup Time Needed:</b> 15 to 30 Minutes (Automated Setup UI)", td_left), Paragraph("<b>Recommended Client Retainer:</b> $99 – $149 / mo", td_left)]
    ]
    t_meta = Table(meta_table_data, colWidths=[270, 234])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 6))
    
    # ── SECTION 1: REALISTIC SETUP BREAKDOWN ──
    elements.append(Paragraph("1. REALISTIC SETUP EFFORT & ONBOARDING CHARGES (15–30 MINS)", h1_style))
    elements.append(Paragraph(
        "Because the AI extraction prompts and regex mappings are already built into the codebase, setting up a new company takes only 15 to 30 minutes via the Setup UI:",
        body_style
    ))
    
    setup_pdf_data = [
        [
            Paragraph("Setup Step", th_style),
            Paragraph("Actual Time", th_style),
            Paragraph("Direct Fee", th_style),
            Paragraph("Why It's Fast & Automated", th_style),
            Paragraph("Fair Setup Fee", th_style),
        ],
        [
            Paragraph("<b>1. Connect Xero OAuth2</b>", td_left),
            Paragraph("5 - 10 min", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("1-click OAuth button in Setup UI (auto-detects tenant ID & base currency)", td_left),
            Paragraph("<b>$100.00</b>", td_right),
        ],
        [
            Paragraph("<b>2. Connect IMAP Mailbox</b>", td_left),
            Paragraph("5 - 10 min", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("Generate 16-char Gmail/M365 App Password, enter into Setup page, test connection", td_left),
            Paragraph("<b>$75.00</b>", td_right),
        ],
        [
            Paragraph("<b>3. Free Gemini API Key</b>", td_left),
            Paragraph("3 - 5 min", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("Generate free key on aistudio.google.com (15 RPM free). No credit card needed", td_left),
            Paragraph("<b>$50.00</b>", td_right),
        ],
        [
            Paragraph("<b>4. Test Ingestion & Verify</b>", td_left),
            Paragraph("2 - 5 min", td_center),
            Paragraph("$0.00", td_right),
            Paragraph("Send 1 sample vendor PDF, verify draft bill creation in live dashboard", td_left),
            Paragraph("<b>$25.00</b>", td_right),
        ],
        [
            Paragraph("<b>TOTAL SETUP PACKAGE</b>", td_bold),
            Paragraph("<b>15 - 30 min</b>", td_center),
            Paragraph("<b>$0.00</b>", td_right),
            Paragraph("<b>Turnkey connection to existing pre-built AI automation</b>", td_bold),
            Paragraph("<b>$250.00 (or $0 promo)</b>", td_bold),
        ]
    ]
    t_setup = Table(setup_pdf_data, colWidths=[120, 55, 50, 204, 75])
    t_setup.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_setup)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 2: MONTHLY CLIENT PRICING PACKAGES ──
    elements.append(Paragraph("2. COMPETITIVE & HIGH-CONVERSION CLIENT PACKAGES", h1_style))
    elements.append(Paragraph(
        "Accessible, high-margin monthly retainers structured for small & medium businesses:",
        body_style
    ))
    
    pkg_pdf_data = [
        [
            Paragraph("Package Tier", th_style),
            Paragraph("Setup Fee", th_style),
            Paragraph("Monthly Retainer", th_style),
            Paragraph("Included Volume", th_style),
            Paragraph("Net Margin", th_style),
            Paragraph("Target Client", th_style),
        ],
        [
            Paragraph("<b>Starter SME Plan</b><br/><i>(Budget friendly)</i>", td_left),
            Paragraph("$150.00", td_right),
            Paragraph("$99.00 / mo", td_right),
            Paragraph("Up to 150 inv/mo", td_center),
            Paragraph("<b>$96.00 / mo (97%)</b>", td_right),
            Paragraph("Small business with 20–150 monthly bills", td_left),
        ],
        [
            Paragraph("<b>Standard Plan</b><br/><i>(Recommended)</i>", td_left),
            Paragraph("$250.00", td_right),
            Paragraph("$149.00 / mo", td_right),
            Paragraph("Up to 500 inv/mo", td_center),
            Paragraph("<b>$144.00 / mo (96%)</b>", td_right),
            Paragraph("Growing business with regular daily invoices", td_left),
        ],
        [
            Paragraph("<b>Pro / High-Volume</b>", td_left),
            Paragraph("$350.00", td_right),
            Paragraph("$249.00 / mo", td_right),
            Paragraph("Up to 1,500 inv/mo", td_center),
            Paragraph("<b>$239.00 / mo (96%)</b>", td_right),
            Paragraph("Trading, e-commerce, multi-entity", td_left),
        ]
    ]
    t_pkg = Table(pkg_pdf_data, colWidths=[110, 65, 75, 75, 84, 95])
    t_pkg.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (4,1), (4,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_pkg)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 3: GEMINI API ZERO HIDDEN COST GUARANTEE ──
    elements.append(Paragraph("3. GEMINI API & ZERO HIDDEN COSTS CLARIFICATION", h1_style))
    elements.append(Paragraph(
        "• <b>100% Free Google AI Tier:</b> Google AI Studio provides 15 Requests Per Minute (RPM) and 1,500 requests/day for <b>$0.00</b>.<br/>"
        "• <b>No Payment Method Required:</b> The client never needs to input a credit card into Google Cloud, eliminating all risk of surprise Google bills.<br/>"
        "• <b>Built-in App Throttling:</b> The codebase strictly rate-limits requests to 15 RPM per tenant, preventing any quota overruns.",
        body_style
    ))
    elements.append(Spacer(1, 6))
    
    # ── SECTION 4: CLIENT ROI ──
    elements.append(Paragraph("4. CLIENT ROI JUSTIFICATION (WHY CLIENTS SAY YES)", h1_style))
    
    roi_pdf_data = [
        [
            Paragraph("Parameter", th_style),
            Paragraph("Manual Data Entry", th_style),
            Paragraph("Automated System", th_style),
            Paragraph("Net Client Benefit", th_style),
        ],
        [
            Paragraph("<b>Monthly Invoice Volume</b>", td_left),
            Paragraph("300 invoices", td_center),
            Paragraph("300 invoices", td_center),
            Paragraph("100% automated intake", td_center),
        ],
        [
            Paragraph("<b>Monthly Bookkeeper Hours</b>", td_left),
            Paragraph("25.0 Hours (@ 5 min/inv)", td_center),
            Paragraph("2.5 Hours (30s review)", td_center),
            Paragraph("<b>22.5 Hours saved / mo</b>", td_center),
        ],
        [
            Paragraph("<b>Monthly Human Labor Cost (@ $25/hr)</b>", td_left),
            Paragraph("$625.00", td_right),
            Paragraph("$62.50", td_right),
            Paragraph("<b>$562.50 / month saved</b>", td_right),
        ],
        [
            Paragraph("<b>Monthly Automation Fee (Standard)</b>", td_left),
            Paragraph("$0.00", td_right),
            Paragraph("$149.00", td_right),
            Paragraph("-$149.00 / month fee", td_right),
        ],
        [
            Paragraph("<b>NET MONTHLY CLIENT BENEFIT</b>", td_bold),
            Paragraph("$625.00", td_bold),
            Paragraph("$211.50", td_bold),
            Paragraph("<b>+$413.50 / MONTH SAVINGS</b>", td_bold),
        ],
        [
            Paragraph("<b>ANNUAL VALUE CREATED</b>", td_bold),
            Paragraph("$7,500.00 / yr", td_bold),
            Paragraph("$2,538.00 / yr", td_bold),
            Paragraph("<b>+$4,962.00 / YEAR NET SAVINGS</b>", td_bold),
        ]
    ]
    t_roi = Table(roi_pdf_data, colWidths=[140, 110, 110, 144])
    t_roi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-3), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_roi)
    elements.append(Spacer(1, 8))
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    print(f"Realistic PDF report created at: {out_path}")


# ==============================================================================
# 3. MARKDOWN REPORT GENERATOR (Realistic Pricing & Zero Hidden Costs)
# ==============================================================================
def create_md_report(out_path):
    md_content = """# Xero Invoice Automation — Realistic Cost Estimation & Client Pricing

**Document Purpose:** SME-Friendly Cost Estimation, Real Setup Effort (15–30 Mins) & Transparent Client Pricing  
**Hardware Baseline:** 8 GB RAM | 160–200 GB NVMe SSD | 4 vCPU  
**Gemini AI Tier:** 100% Free Google AI Studio Tier (15 RPM / 1,500 req/day — $0.00)  

---

## 1. Transparent Setup Effort Audit (Why It Takes Only 15–30 Mins)

Because the prompt engineering, PDF parsing pipeline, and regex fallback mapping are **already fully built into your codebase**, setting up a new company requires zero coding:

| Setup Step | Actual Time Needed | Direct Vendor Fee | Why It Is Fast & Automated | Recommended Client Setup Charge |
|---|---|---|---|---|
| **1. Connect Xero OAuth2** | 5 – 10 mins | **$0.00** | 1-click OAuth button in Setup UI (auto-detects tenant ID & base currency). | **$100.00** |
| **2. Connect IMAP Mailbox** | 5 – 10 mins | **$0.00** | Generate a 16-character Gmail/M365 App Password and paste into Setup page. | **$75.00** |
| **3. Free Gemini API Key** | 3 – 5 mins | **$0.00** | Generate free API key on aistudio.google.com (15 RPM free, no credit card). | **$50.00** |
| **4. Test Ingestion & Verify** | 2 – 5 mins | **$0.00** | Send 1 sample vendor PDF, check live dashboard for draft creation in Xero. | **$25.00** |
| **TOTAL SETUP EFFORT** | **15 – 30 mins** | **$0.00** | **Turnkey connection to existing pre-built pipeline** | **$250.00 (or $0 on annual promo)** |

---

## 2. Gemini API: 100% Free with Zero Hidden Costs for Client

* **Google AI Studio Free Tier:** Gives **15 Requests Per Minute (RPM)** and **1,500 requests per day** for **$0.00**.
* **Zero Billing / Zero Credit Card Risk:** The client does not need to activate a billing account or add a credit card on Google Cloud.
* **Built-in App Throttling:** Your app's `gemini-client.js` enforces a 15 RPM sliding window rate limit per tenant and rotates models (`gemini-3.5-flash-lite` → `gemini-3.1-flash-lite`), so the client will never incur surprise Google charges.

---

## 3. Practical & Competitive Client Packages (SME Friendly)

| Package Tier | One-Time Setup Fee | Monthly Retainer | Included Volume | Real Server Cost | Net Monthly Margin | Target Client Profile |
|---|---|---|---|---|---|---|
| **Starter SME Plan (Budget)** | **$150.00** | **$99.00 / mo** | Up to 150 inv/mo | ~$3.00 / mo | **$96.00 / mo (97%)** | Small business with 20–150 monthly bills |
| **Standard Plan (Recommended)** | **$250.00** | **$149.00 / mo** | Up to 500 inv/mo | ~$5.00 / mo | **$144.00 / mo (96%)** | Standard SME with daily supplier bills |
| **Pro / High-Volume Plan** | **$350.00** | **$249.00 / mo** | Up to 1,500 inv/mo | ~$10.00 / mo | **$239.00 / mo (96%)** | Trading, e-commerce, multi-entity |

---

## 4. Multi-Company SaaS Profitability on 8GB Server ($55.25/mo)

Because your **8GB DigitalOcean server ($55.25/mo)** easily handles 20 to 50 connected companies:

| Connected Clients | Average Retainer | Gross Monthly Revenue | Total Server Cost | Net Monthly Profit | Gross Margin |
|---|---|---|---|---|---|
| **5 Clients** | $149.00 / mo | **$745.00 / mo** | $55.25 / mo | **+$689.75 / month** | **92.6%** |
| **15 Clients** | $149.00 / mo | **$2,235.00 / mo** | $55.25 / mo | **+$2,179.75 / month** | **97.5%** |
| **35 Clients** | $149.00 / mo | **$5,215.00 / mo** | $55.25 / mo | **+$5,159.75 / month** | **98.9%** |

---

## 5. Client ROI Justification (Why Clients Say Yes Immediately)

* **Manual Bookkeeper Cost:** 300 invoices/month = **25 hours of typing** @ $25/hr = **$625.00 / month**.
* **Automated Cost (Standard Plan):** 2.5 hours review ($62.50) + $149 retainer = **$211.50 / month total**.
* **Net Monthly Savings to Client:** **+$413.50 / month ($4,962.00 / year)**.
* **Payback Period:** Less than **1 month**.
"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Realistic Markdown report created at: {out_path}")


if __name__ == "__main__":
    # Generate in root /report
    create_excel_report(os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.xlsx"))
    create_pdf_report(os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.pdf"))
    create_md_report(os.path.join(REPORT_DIR, "COST_ESTIMATION_AND_PRICING_REPORT.md"))

    # Generate in /report/arch_recommendation
    create_excel_report(os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.xlsx"))
    create_pdf_report(os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.pdf"))
    create_md_report(os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.md"))

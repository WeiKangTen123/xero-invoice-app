import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

REPORT_DIR = "/Users/weikangten/Desktop/xero-invoice-app-master/report"
ARCH_DIR = os.path.join(REPORT_DIR, "arch_recommendation")

PDF_PATH = os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.pdf")
XLSX_PATH = os.path.join(REPORT_DIR, "XERO_AUTOMATION_COST_ESTIMATION_REPORT.xlsx")
MD_PATH = os.path.join(REPORT_DIR, "COST_ESTIMATION_AND_PRICING_REPORT.md")

ARCH_PDF_PATH = os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.pdf")
ARCH_XLSX_PATH = os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.xlsx")
ARCH_MD_PATH = os.path.join(ARCH_DIR, "SYSTEM_ARCHITECTURE_AND_COST_RECOMMENDATION.md")

# ==============================================================================
# 1. REPORTLAB NUMBERED CANVAS (Professional Headers & Footers)
# ==============================================================================
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        # Header (pages 2+)
        if self._pageNumber > 1:
            self.drawString(54, 11 * inch - 36, "XERO INVOICE AUTOMATION — COST ESTIMATION & SERVICE PROPOSAL")
            self.drawRightString(8.5 * inch - 54, 11 * inch - 36, "CONFIDENTIAL")
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(54, 11 * inch - 42, 8.5 * inch - 54, 11 * inch - 42)
            
        # Footer (all pages)
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(54, 46, 8.5 * inch - 54, 46)
        
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawString(54, 32, "Prepared for Client Commercial Quotation | August 2026")
        self.drawRightString(8.5 * inch - 54, 32, page_str)
        self.restoreState()


# ==============================================================================
# 2. PDF REPORT GENERATOR (Clean Client-Facing Presentation)
# ==============================================================================
def create_pdf(out_path):
    doc = SimpleDocTemplate(
        out_path,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    c_navy = colors.HexColor("#1E3A8A")
    c_slate = colors.HexColor("#334155")
    c_body = colors.HexColor("#1E293B")
    c_teal = colors.HexColor("#0D9488")
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=18, leading=22,
        textColor=c_navy, spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'],
        fontName='Helvetica-Oblique', fontSize=9.5, leading=13,
        textColor=colors.HexColor("#64748B"), spaceAfter=10
    )
    h1_style = ParagraphStyle(
        'SecH1', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10.5, leading=14,
        textColor=c_navy, spaceBefore=9, spaceAfter=4, keepWithNext=True
    )
    body_style = ParagraphStyle(
        'BodyDark', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, leading=11,
        textColor=c_body, spaceAfter=4
    )
    th_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9.5,
        textColor=colors.white, alignment=1
    )
    td_left = ParagraphStyle(
        'TableCellLeft', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_bold = ParagraphStyle(
        'TableCellBold', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.2, leading=9,
        textColor=c_body
    )
    td_right = ParagraphStyle(
        'TableCellRight', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=2
    )
    td_center = ParagraphStyle(
        'TableCellCenter', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7.2, leading=9,
        textColor=c_body, alignment=1
    )
    
    elements = []
    
    # ── HEADER BLOCK ──
    elements.append(Paragraph("XERO INVOICE AUTOMATION — COST ESTIMATION REPORT", title_style))
    elements.append(Paragraph("Commercial Quotation, Managed Server Hosting, Maintenance SLA & Client ROI Model", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=c_navy, spaceAfter=8))
    
    # ── METADATA TABLE ──
    meta_table_data = [
        [Paragraph("<b>Target System:</b> Turnkey Email-to-Xero Automation", td_left), Paragraph("<b>Prepared Date:</b> August 2026", td_left)],
        [Paragraph("<b>Service Model:</b> Fully Managed Cloud Server + Maintenance SLA", td_left), Paragraph("<b>Base Currency:</b> USD ($)", td_left)],
        [Paragraph("<b>Setup Scope:</b> 1-Click Xero OAuth + IMAP Mailbox + Verification", td_left), Paragraph("<b>Setup Turnaround:</b> Same-Day Onboarding (15–30 Mins)", td_left)]
    ]
    t_meta = Table(meta_table_data, colWidths=[270, 234])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 6))
    
    # ── SECTION 1: MASTER COMMERCIAL PACKAGES ──
    elements.append(Paragraph("1. RECOMMENDED COMMERCIAL PACKAGES (SME PRICING)", h1_style))
    elements.append(Paragraph(
        "Turnkey packages combining one-time onboarding, 24/7 cloud server hosting, automated AI extraction, and ongoing maintenance SLA:",
        body_style
    ))
    
    pkg_pdf_data = [
        [
            Paragraph("Package Tier", th_style),
            Paragraph("One-Time Setup", th_style),
            Paragraph("Monthly Retainer", th_style),
            Paragraph("Monthly Invoice Quota", th_style),
            Paragraph("Included Services & SLA", th_style),
        ],
        [
            Paragraph("<b>Starter SME Plan</b><br/><i>(For small business)</i>", td_left),
            Paragraph("$150.00", td_right),
            Paragraph("<b>$100.00 / mo</b>", td_right),
            Paragraph("Up to 200 invoices / mo", td_center),
            Paragraph("Cloud server hosting, real-time IMAP watcher, standard email support, automated daily backups", td_left),
        ],
        [
            Paragraph("<b>Standard Business</b><br/><i>(Recommended — Best Value)</i>", td_left),
            Paragraph("$250.00", td_right),
            Paragraph("<b>$150.00 / mo</b>", td_right),
            Paragraph("Up to 600 invoices / mo", td_center),
            Paragraph("High-speed server hosting, priority parsing queue, Xero OAuth maintenance, uptime monitoring SLA", td_left),
        ],
        [
            Paragraph("<b>Pro / High-Volume</b><br/><i>(High volume / Multi-entity)</i>", td_left),
            Paragraph("$350.00", td_right),
            Paragraph("<b>$200.00 / mo</b>", td_right),
            Paragraph("Up to 1,500 invoices / mo", td_center),
            Paragraph("Dedicated high-volume queue, multi-currency support, custom account codes, priority on-call support", td_left),
        ]
    ]
    t_pkg = Table(pkg_pdf_data, colWidths=[110, 65, 75, 95, 159])
    t_pkg.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (2,2), (2,2), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_pkg)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 2: ITEMIZED COST BREAKDOWN (SETUP & MONTHLY) ──
    elements.append(Paragraph("2. ITEMIZED SERVICE & COST BREAKDOWN", h1_style))
    elements.append(Paragraph(
        "A clear, transparent breakdown of onboarding, hosting, maintenance, and third-party software fees:",
        body_style
    ))
    
    item_pdf_data = [
        [
            Paragraph("Component / Service Item", th_style),
            Paragraph("Billing Frequency", th_style),
            Paragraph("Fee (USD)", th_style),
            Paragraph("Scope & Description", th_style),
        ],
        [
            Paragraph("<b>1. System Onboarding & Setup</b>", td_left),
            Paragraph("One-Time Fee", td_center),
            Paragraph("<b>$250.00</b>", td_right),
            Paragraph("Xero OAuth2 connection, IMAP email mailbox linking, chart of accounts mapping (Account 310 / GST rates), and pilot verification test.", td_left),
        ],
        [
            Paragraph("<b>2. Cloud Server Hosting & Pipeline</b>", td_left),
            Paragraph("Monthly Recurring", td_center),
            Paragraph("<b>$100.00 / mo</b>", td_right),
            Paragraph("24/7 dedicated cloud server VPS (8GB RAM / 200GB SSD), real-time email watcher, automated AI extraction queue, and live review dashboard.", td_left),
        ],
        [
            Paragraph("<b>3. System Maintenance & Support SLA</b>", td_left),
            Paragraph("Monthly Recurring", td_center),
            Paragraph("<b>$50.00 / mo</b>", td_right),
            Paragraph("Continuous uptime monitoring, Xero OAuth token auto-refreshes, daily SQLite database backups, error triage, and technical maintenance.", td_left),
        ],
        [
            Paragraph("<b>TOTAL MONTHLY SERVICE RETAINER</b>", td_bold),
            Paragraph("<b>Monthly (Bundled)</b>", td_center),
            Paragraph("<b>$150.00 / mo</b>", td_bold),
            Paragraph("<b>Complete all-inclusive managed software service (Standard Tier)</b>", td_bold),
        ],
        [
            Paragraph("<b>4. Client Xero Subscription (Standard)</b>", td_left),
            Paragraph("Monthly (Direct to Xero)", td_center),
            Paragraph("~$35 – $45 / mo", td_right),
            Paragraph("Client's existing Xero business plan (e.g. Grow / Growing plan for unlimited bills). Paid directly to Xero.", td_left),
        ],
        [
            Paragraph("<b>5. Email Inbox Connection (IMAP)</b>", td_left),
            Paragraph("Monthly", td_center),
            Paragraph("$0.00 (Included)", td_right),
            Paragraph("Connects directly to client's existing mailbox (Gmail / Google Workspace / Microsoft 365 / Outlook).", td_left),
        ]
    ]
    t_item = Table(item_pdf_data, colWidths=[140, 80, 75, 209])
    t_item.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-3), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,4), (-1,4), colors.HexColor("#DBEAFE")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(t_item)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 3: WHAT THE MAINTENANCE FEE COVERS ──
    elements.append(Paragraph("3. WHAT IS INCLUDED IN THE ONGOING MAINTENANCE SLA ($50/MO)", h1_style))
    elements.append(Paragraph(
        "The ongoing maintenance retainer ensures the automation runs smoothly 24/7 without requiring any internal IT staff on the client's side:",
        body_style
    ))
    
    maint_pdf_data = [
        [
            Paragraph("Maintenance Scope", th_style),
            Paragraph("What We Manage For You", th_style),
            Paragraph("Client Value & Benefit", th_style),
        ],
        [
            Paragraph("<b>24/7 Uptime & Watcher Health</b>", td_left),
            Paragraph("Monitoring persistent IMAP connections to ensure no incoming supplier email is missed or delayed.", td_left),
            Paragraph("Zero lost bills, 100% reliable real-time intake.", td_left),
        ],
        [
            Paragraph("<b>Xero OAuth Token Maintenance</b>", td_left),
            Paragraph("Managing token lifecycle, auto-refreshes, and preventing authentication disconnects.", td_left),
            Paragraph("Uninterrupted sync into Xero without re-login friction.", td_left),
        ],
        [
            Paragraph("<b>Daily Automated DB Backups</b>", td_left),
            Paragraph("Automated daily snapshots of all invoice records, transaction histories, and audit trails.", td_left),
            Paragraph("Complete disaster recovery protection & tax audit safety.", td_left),
        ],
        [
            Paragraph("<b>Parsing Support & Bug Fixes</b>", td_left),
            Paragraph("Fast support if an unusual supplier PDF format requires adjustment or verification.", td_left),
            Paragraph("Dedicated technical resolution within guaranteed SLA.", td_left),
        ]
    ]
    t_maint = Table(maint_pdf_data, colWidths=[120, 214, 170])
    t_maint.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_navy),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_maint)
    elements.append(Spacer(1, 8))
    
    elements.append(PageBreak())
    
    # ── SECTION 4: CLIENT RETURN ON INVESTMENT ──
    elements.append(Paragraph("4. CLIENT RETURN ON INVESTMENT (ROI) & LABOR SAVINGS", h1_style))
    elements.append(Paragraph(
        "Financial model proving the immediate cost reduction achieved by replacing manual data entry with the automated pipeline:",
        body_style
    ))
    
    roi_pdf_data = [
        [
            Paragraph("Parameter / Financial Metric", th_style),
            Paragraph("Manual Data Entry", th_style),
            Paragraph("Automated Pipeline (Standard)", th_style),
            Paragraph("Net Client Benefit", th_style),
        ],
        [
            Paragraph("<b>Monthly Invoice Volume</b>", td_left),
            Paragraph("300 invoices / mo", td_center),
            Paragraph("300 invoices / mo", td_center),
            Paragraph("100% automated intake", td_center),
        ],
        [
            Paragraph("<b>Human Labor Required</b>", td_left),
            Paragraph("25.0 Hours (@ 5 min/inv)", td_center),
            Paragraph("2.5 Hours (30s review)", td_center),
            Paragraph("<b>22.5 Hours saved / month</b>", td_center),
        ],
        [
            Paragraph("<b>Monthly Labor Cost (@ $25/hr)</b>", td_left),
            Paragraph("$625.00 / month", td_right),
            Paragraph("$62.50 / month", td_right),
            Paragraph("<b>$562.50 / month saved</b>", td_right),
        ],
        [
            Paragraph("<b>Managed Server & Maintenance Fee</b>", td_left),
            Paragraph("$0.00", td_right),
            Paragraph("$150.00 / month", td_right),
            Paragraph("-$150.00 / month fee", td_right),
        ],
        [
            Paragraph("<b>Xero Business Subscription (Baseline)</b>", td_left),
            Paragraph("$40.00 / month", td_right),
            Paragraph("$40.00 / month", td_right),
            Paragraph("$0.00 (Standard baseline)", td_right),
        ],
        [
            Paragraph("<b>TOTAL MONTHLY COST TO CLIENT</b>", td_bold),
            Paragraph("<b>$665.00 / month</b>", td_bold),
            Paragraph("<b>$252.50 / month</b>", td_bold),
            Paragraph("<b>+$412.50 / MONTH SAVINGS</b>", td_bold),
        ],
        [
            Paragraph("<b>ANNUAL NET VALUE TO CLIENT</b>", td_bold),
            Paragraph("<b>$7,980.00 / year</b>", td_bold),
            Paragraph("<b>$3,030.00 / year</b>", td_bold),
            Paragraph("<b>+$4,950.00 / YEAR NET PROFIT</b>", td_bold),
        ]
    ]
    t_roi = Table(roi_pdf_data, colWidths=[140, 110, 110, 144])
    t_roi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_slate),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-3), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor("#CCFBF1")),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
    ]))
    elements.append(t_roi)
    elements.append(Spacer(1, 8))
    
    # ── SECTION 5: EXECUTIVE SUMMARY ──
    elements.append(Paragraph("5. EXECUTIVE SUMMARY FOR CLIENT QUOTATION", h1_style))
    elements.append(Paragraph(
        "• <b>One-Time Setup Fee ($250):</b> Covers turnkey connection of Xero OAuth, IMAP email mailbox, Chart of Accounts mapping, and pilot test verification.<br/>"
        "• <b>Monthly Managed Retainer ($150/mo):</b> Combines dedicated 24/7 cloud server hosting ($100/mo) and ongoing system maintenance & uptime support SLA ($50/mo).<br/>"
        "• <b>Client ROI:</b> The client saves <b>$412.50 every month ($4,950 / year)</b> in manual bookkeeping labor, paying off the setup fee in <b>less than 3 weeks</b>.",
        body_style
    ))
    
    doc.build(elements, canvasmaker=NumberedCanvas)
    print(f"Client PDF successfully generated at: {out_path}")


# ==============================================================================
# 3. EXCEL WORKBOOK GENERATOR (Matching Client Presentation)
# ==============================================================================
def create_excel(out_path):
    wb = openpyxl.Workbook()
    
    NAVY_DARK = "1E3A8A"
    NAVY_LIGHT = "DBEAFE"
    SLATE_DARK = "334155"
    SLATE_LIGHT = "F8FAFC"
    GREEN_LIGHT = "CCFBF1"
    GRAY_TEXT = "64748B"
    WHITE = "FFFFFF"
    
    font_title = Font(name="Calibri", size=16, bold=True, color=NAVY_DARK)
    font_subtitle = Font(name="Calibri", size=11, italic=True, color=GRAY_TEXT)
    font_sec_hdr = Font(name="Calibri", size=12, bold=True, color=WHITE)
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=11, color="1E293B")
    font_data_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_total = Font(name="Calibri", size=11, bold=True, color=NAVY_DARK)
    
    fill_sec_hdr = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=SLATE_DARK, end_color=SLATE_DARK, fill_type="solid")
    fill_alt_row = PatternFill(start_color=SLATE_LIGHT, end_color=SLATE_LIGHT, fill_type="solid")
    fill_highlight = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    fill_total = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    total_border = Border(
        top=Side(style='thin', color='1E3A8A'),
        bottom=Side(style='double', color='1E3A8A'),
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1')
    )
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    # Tab 1: Commercial Proposal
    ws1 = wb.active
    ws1.title = "Commercial Proposal"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "XERO INVOICE AUTOMATION — CLIENT COST ESTIMATION & QUOTATION"
    ws1["A1"].font = font_title
    
    ws1.merge_cells("A2:F2")
    ws1["A2"] = "Turnkey Service Model: One-Time Onboarding, Managed Cloud Server, Maintenance SLA & ROI"
    ws1["A2"].font = font_subtitle
    
    ws1["A4"] = "Service Scope:"
    ws1["B4"] = "Turnkey Email-to-Xero Sync Pipeline"
    ws1["A5"] = "Prepared Date:"
    ws1["B5"] = "August 2026"
    ws1["D4"] = "Currency:"
    ws1["E4"] = "USD ($)"
    ws1["D5"] = "Server Hardware:"
    ws1["E5"] = "8 GB RAM | 200 GB NVMe SSD"
    for r in [4, 5]:
        ws1[f"A{r}"].font = font_data_bold
        ws1[f"B{r}"].font = font_data
        ws1[f"D{r}"].font = font_data_bold
        ws1[f"E{r}"].font = font_data

    # Section 1: Packages Table
    ws1.merge_cells("A7:F7")
    ws1["A7"] = "RECOMMENDED COMMERCIAL PACKAGES"
    ws1["A7"].font = font_sec_hdr
    ws1["A7"].fill = fill_sec_hdr

    headers_pkg = ["Package Tier", "One-Time Setup Fee", "Monthly Server & Maint.", "Included Volume", "Net Monthly Margin", "Target Client Profile"]
    for c_idx, h in enumerate(headers_pkg, 1):
        cell = ws1.cell(row=8, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    pkg_data = [
        ["Starter SME Plan", 150.00, 100.00, "Up to 200 invoices/mo", 95.00, "Small business with 20–200 monthly bills"],
        ["Standard Business Plan (Recommended)", 250.00, 150.00, "Up to 600 invoices/mo", 145.00, "Standard SME with daily supplier invoices"],
        ["Pro / High-Volume Plan", 350.00, 200.00, "Up to 1,500 invoices/mo", 190.00, "Trading, logistics, e-commerce with heavy volume"],
    ]

    for r_idx, row_vals in enumerate(pkg_data, start=9):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx in [2, 3, 5]:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
                if c_idx == 5:
                    cell.fill = fill_highlight
                    cell.font = font_total
            else:
                cell.font = font_data
                cell.alignment = align_center if c_idx == 4 else align_left

    # Section 2: Itemized Cost Table
    ws1.merge_cells("A13:F13")
    ws1["A13"] = "ITEMIZED SERVICE & SUBSCRIPTION BREAKDOWN"
    ws1["A13"].font = font_sec_hdr
    ws1["A13"].fill = fill_sec_hdr

    headers_item = ["Service / Subscription Component", "Billing Frequency", "Fee (USD)", "Payer / Vendor", "Description & Inclusions"]
    for c_idx, h in enumerate(headers_item, 1):
        cell = ws1.cell(row=14, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center

    item_data = [
        ["1. System Onboarding & Setup", "One-Time Fee", 250.00, "Billed by You", "Turnkey Xero OAuth, IMAP Mailbox, CoA Account 310 & GST mapping"],
        ["2. Managed Cloud Server Hosting", "Monthly Recurring", 100.00, "Billed by You", "24/7 dedicated 8GB cloud server, real-time IMAP watcher, review dashboard"],
        ["3. System Maintenance & Support SLA", "Monthly Recurring", 50.00, "Billed by You", "24/7 uptime monitoring, Xero token auto-refresh, daily backups, support"],
        ["4. Client Xero Subscription (Grow Plan)", "Monthly Recurring", 40.00, "Direct to Xero", "Client's existing Xero business software plan (unlimited bills)"],
        ["5. Email Inbox Connection (IMAP)", "Monthly", 0.00, "Client Mailbox", "Connects to existing Gmail / Google Workspace / Microsoft 365 inbox"]
    ]

    for r_idx, row_vals in enumerate(item_data, start=15):
        fill_to_use = fill_alt_row if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if fill_to_use.fill_type:
                cell.fill = fill_to_use
            if c_idx == 1:
                cell.font = font_data_bold
                cell.alignment = align_left
            elif c_idx == 3:
                cell.font = font_data
                cell.alignment = align_right
                cell.number_format = "$#,##0.00"
            elif c_idx in [2, 4]:
                cell.font = font_data
                cell.alignment = align_center
            else:
                cell.font = font_data
                cell.alignment = align_left

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str += "    "
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
    wb.save(out_path)
    print(f"Excel report successfully generated at: {out_path}")


# ==============================================================================
# 4. MARKDOWN GENERATOR
# ==============================================================================
def create_md(out_path):
    md_content = """# Xero Invoice Automation — Cost Estimation & Service Proposal

**Document Purpose:** Commercial Proposal, Managed Server Hosting, Maintenance SLA & Client ROI  
**Target Hardware:** Dedicated 8 GB RAM | 200 GB NVMe SSD  
**Architecture:** Turnkey Multi-User Real-Time Email-to-Xero Sync Pipeline  

---

## 1. Commercial Pricing Packages (SME Friendly)

| Package Tier | One-Time Setup Fee | Monthly Server & Maintenance | Included Monthly Invoices | Included Services & SLA |
|---|---|---|---|---|
| **Starter SME Plan** | **$150.00** | **$100.00 / mo** | Up to 200 invoices / mo | Dedicated cloud server, real-time IMAP watcher, daily backups, standard email support. |
| **Standard Business (Recommended)** | **$250.00** | **$150.00 / mo** | Up to 600 invoices / mo | High-speed server, priority queue, Xero OAuth maintenance, 24/7 uptime monitoring SLA. |
| **Pro / High-Volume Plan** | **$350.00** | **$200.00 / mo** | Up to 1,500 invoices / mo | High-volume queue, multi-currency support, custom account codes, priority support. |

---

## 2. Itemized Service & Subscription Breakdown

| Component / Service Item | Billing Frequency | Fee (USD) | Billed By | Scope & Deliverables |
|---|---|---|---|---|
| **1. System Onboarding & Setup** | One-Time Fee | **$250.00** | Billed by You | Xero OAuth connection, IMAP mailbox linking, Chart of Accounts mapping (Account 310 / GST rates), and pilot test verification. |
| **2. Managed Cloud Server Hosting** | Monthly Recurring | **$100.00 / mo** | Billed by You | 24/7 dedicated 8GB cloud server, real-time email watcher, automated extraction queue, and live review dashboard. |
| **3. System Maintenance & Support SLA** | Monthly Recurring | **$50.00 / mo** | Billed by You | Continuous uptime monitoring, Xero OAuth token auto-refreshes, daily SQLite database backups, error triage, and technical support. |
| **TOTAL MONTHLY SERVICE RETAINER** | **Monthly (Bundled)** | **$150.00 / mo** | **Billed by You** | **Complete all-inclusive managed software service (Standard Tier)** |
| **4. Client Xero Subscription (Grow Plan)** | Monthly Recurring | **~$35 – $45 / mo** | Direct to Xero | Client's existing Xero business plan (for unlimited bills). Paid directly to Xero. |
| **5. Email Inbox Connection (IMAP)** | Monthly | **$0.00** | Client Mailbox | Connects directly to client's existing Gmail / Google Workspace / Microsoft 365 inbox. |

---

## 3. What the Monthly Maintenance Fee Covers ($50.00 / mo)

The ongoing maintenance fee ensures your automation runs continuously without requiring internal IT staff:
1. **24/7 Uptime & Watcher Health:** Monitoring persistent IMAP connections to guarantee zero dropped emails or delayed supplier bills.
2. **Xero OAuth Token Auto-Refreshes:** Proactively managing token lifecycles to prevent authentication expirations.
3. **Daily Automated DB Backups:** Automated daily snapshots of all invoice records, transaction histories, and audit logs.
4. **Parsing Support & Bug Fixes:** Dedicated technical resolution if an unusual supplier invoice layout requires verification.

---

## 4. Client Return on Investment (ROI) & Labor Savings

| Financial Parameter | Manual Data Entry | Automated Pipeline (Standard Tier) | Net Client Savings |
|---|---|---|---|
| **Monthly Invoice Volume** | 300 invoices / mo | 300 invoices / mo | 100% automated intake |
| **Staff Labor Time Required** | 25.0 Hours (@ 5 min/inv) | 2.5 Hours (30s review) | **22.5 Hours saved / month** |
| **Human Labor Expense (@ $25/hr)** | **$625.00 / month** | $62.50 / month | **$562.50 / month saved** |
| **Managed Server & Maintenance Fee** | $0.00 | **$150.00 / month** | -$150.00 / month fee |
| **Xero Business Subscription (Baseline)** | $40.00 / month | $40.00 / month | $0.00 (Standard baseline) |
| **TOTAL MONTHLY COST TO CLIENT** | **$665.00 / month** | **$252.50 / month** | **+$412.50 / MONTH NET SAVINGS** |
| **ANNUAL NET VALUE CREATED** | **$7,980.00 / year** | **$3,030.00 / year** | **+$4,950.00 / YEAR NET PROFIT** |

---

## 5. Executive Summary for Client Proposal

* **One-Time Setup Fee ($250):** Covers turnkey connection of Xero OAuth, IMAP email mailbox, Chart of Accounts mapping, and pilot test verification.
* **Monthly Managed Retainer ($150/mo):** Combines dedicated 24/7 cloud server hosting ($100/mo) and ongoing system maintenance & uptime support SLA ($50/mo).
* **Client ROI:** The client saves **$412.50 every month ($4,950 / year)** in manual bookkeeping labor, paying off the setup fee in **less than 3 weeks**.
"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"Markdown report successfully generated at: {out_path}")


if __name__ == "__main__":
    create_pdf(PDF_PATH)
    create_excel(XLSX_PATH)
    create_md(MD_PATH)

    create_pdf(ARCH_PDF_PATH)
    create_excel(ARCH_XLSX_PATH)
    create_md(ARCH_MD_PATH)
